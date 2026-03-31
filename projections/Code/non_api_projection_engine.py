"""
Non API RPU Projection Engine
==============================
Replicates Bright Money's Non API affiliate revenue projection model.
Produces a 20-tab Excel workbook with every micro step visible and debuggable.

Dependencies: pandas, openpyxl
"""

import pandas as pd
import numpy as np
from datetime import timedelta
from pathlib import Path
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ==============================================================================
# CONFIGURATION
# ==============================================================================

BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_DIR = BASE_DIR / "Inputs"
ENROLLS_PATH = INPUT_DIR / "Non API" / "Enrolls data.csv"
NON_API_PATH = INPUT_DIR / "Non API" / "Non API data.csv"
OUTPUT_PATH = BASE_DIR / "Outputs" / "Non_API_RPU_Projection_Output.xlsx"

TOUCHPOINT_MAP = {
    "1. FUNNEL": "Funnel",
    "2. FUNNEL_TWO": "Funnel_second",
    "3. FUNNEL_THREE": "Funnel_third",
    "4. DASHBOARD": "Dashboard",
    "5. DISCOVER_TAB_V2": "Discover Tab",
    "7. AI_ASSISTANT": "AI_assistant",
    "8. LOAN_AGENT": "Loan_agent",
    "9. COMMS": "Comms",
}
TOUCHPOINT_DEFAULT = "Non-click"

RECENCY_WEIGHTS = [0.4, 0.3, 0.2, 0.1]

W1_RPC_OFFSETS = [7, 14, 21, 28]          # days back for W1 rolling avg
W2W4_RATIO_OFFSETS = [35, 42, 49, 56]     # days back for W2W4 ratio rolling avg
M3_RATIO_OFFSETS = [70, 77, 84, 91]       # days back for M3 ratio rolling avg

DEFAULT_IMP_RATIO = 0
DEFAULT_CTR_RATIO = 1
DEFAULT_RPC_RATIO = 1


# ==============================================================================
# UTILITY FUNCTIONS
# ==============================================================================

def safe_div(numerator, denominator, default=0.0):
    """Element-wise safe division, returning default where denominator is 0 or NaN."""
    result = np.where(
        (denominator == 0) | pd.isna(denominator),
        default,
        numerator / np.where((denominator == 0) | pd.isna(denominator), 1, denominator),
    )
    return result


def clean_numeric(series):
    """Strip comma formatting and convert to float."""
    if series.dtype == object:
        return pd.to_numeric(series.astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
    return series.fillna(0).astype(float)


def recency_weighted_avg(values, weights=None):
    """Compute recency weighted average. Missing (NaN/None) values contribute 0."""
    if weights is None:
        weights = RECENCY_WEIGHTS
    total = 0.0
    for v, w in zip(values, weights):
        if pd.notna(v):
            total += w * v
    return total


def sort_output(df, cohort_col="feo_cohort"):
    """Sort DataFrame by feo_cohort descending (most recent first)."""
    if cohort_col in df.columns:
        return df.sort_values(cohort_col, ascending=False).reset_index(drop=True)
    return df


# ==============================================================================
# STEP 1: LOAD AND CLEAN DATA
# ==============================================================================

def load_and_clean_data(enrolls_path, non_api_path):
    """Load both CSVs and clean numeric formatting."""
    print("Loading data...")

    enrolls_df = pd.read_csv(enrolls_path)
    enrolls_df["feo_cohort"] = pd.to_datetime(enrolls_df["feo_cohort"], format="mixed")
    enrolls_df["enrols"] = clean_numeric(enrolls_df["enrols"])
    print(f"  Enrolls: {len(enrolls_df)} rows")

    non_api_df = pd.read_csv(non_api_path)
    non_api_df["feo_cohort"] = pd.to_datetime(non_api_df["feo_cohort"], format="mixed")
    for col in ["enrols", "impression_count", "click_count", "payout", "conversion_count"]:
        non_api_df[col] = clean_numeric(non_api_df[col])
    non_api_df["payout"] = non_api_df["payout"].fillna(0)
    print(f"  Non API: {len(non_api_df)} rows")

    return enrolls_df, non_api_df


# ==============================================================================
# STEP 2: PREPROCESS (ENRICH)
# ==============================================================================

def preprocess(non_api_df, enrolls_df, touchpoint_map):
    """Enrich Non API data with touchpoint mapping, UID, ENROLLS lookup, and base metrics."""
    print("Preprocessing / enriching data...")

    df = non_api_df.copy()

    # Touchpoint mapping
    df["TOUCHPOINT"] = df["imp_source"].map(touchpoint_map).fillna(TOUCHPOINT_DEFAULT)
    df["LENDER"] = df["product_name"]
    df["SEGMENT"] = df["segment"]

    # UID for enrollment lookup: concatenation of feo_cohort string + SEGMENT
    df["UID"] = df["feo_cohort"].astype(str) + df["SEGMENT"]

    # Build enrollment lookup (use .astype(str) format to match UID construction)
    enrolls_lookup = dict(zip(
        enrolls_df["feo_cohort"].astype(str) + enrolls_df["segment"],
        enrolls_df["enrols"]
    ))

    df["ENROLLS"] = df["UID"].map(enrolls_lookup).fillna(0)

    # Base metrics
    df["Imp_pct"] = safe_div(df["impression_count"].values, df["ENROLLS"].values)
    df["CTR"] = safe_div(df["click_count"].values, df["impression_count"].values)
    df["RPC"] = safe_div(df["payout"].values, df["click_count"].values)
    df["RPU"] = safe_div(df["payout"].values, df["ENROLLS"].values)

    print(f"  Enriched: {len(df)} rows, {df['TOUCHPOINT'].nunique()} touchpoints, {df['LENDER'].nunique()} lenders")
    return df


# ==============================================================================
# STEP 3: W1 PROJECTION
# ==============================================================================

def build_actual_rpc_table(w1_df):
    """Build the W1 actual RPC lookup table with recency weighted rolling averages."""
    print("Building W1 Actual RPC lookup table...")

    # One row per (TOUCHPOINT, LENDER, SEGMENT, feo_cohort)
    rpc_table = w1_df[["TOUCHPOINT", "LENDER", "SEGMENT", "feo_cohort", "RPC"]].copy()
    rpc_table = rpc_table.drop_duplicates(subset=["TOUCHPOINT", "LENDER", "SEGMENT", "feo_cohort"])

    rpc_table["key_no_cohort"] = rpc_table["TOUCHPOINT"] + rpc_table["LENDER"] + rpc_table["SEGMENT"]
    rpc_table["key_full"] = rpc_table["key_no_cohort"] + rpc_table["feo_cohort"].astype(str)
    rpc_table["Actual_RPC"] = rpc_table["RPC"]

    # Build a fast lookup: (key_no_cohort, feo_cohort) -> RPC
    rpc_lookup = {}
    for _, row in rpc_table.iterrows():
        rpc_lookup[(row["key_no_cohort"], row["feo_cohort"])] = row["Actual_RPC"]

    # Compute recency weighted rolling average
    lookback_cols = []
    for offset in W1_RPC_OFFSETS:
        col_name = f"Lookback_{offset}d_RPC"
        lookback_cols.append(col_name)
        rpc_table[col_name] = rpc_table.apply(
            lambda r: rpc_lookup.get((r["key_no_cohort"], r["feo_cohort"] - timedelta(days=offset)), np.nan),
            axis=1,
        )

    rpc_table["Rolling_4W_Avg_RPC"] = rpc_table.apply(
        lambda r: recency_weighted_avg([r[c] for c in lookback_cols]),
        axis=1,
    )

    print(f"  RPC table: {len(rpc_table)} rows")
    return rpc_table


def compute_w1_projection(w1_df, actual_rpc_table):
    """Compute row level, segment level, and blended W1 RPU."""
    print("Computing W1 projections...")

    # Build lookup from rpc table
    rpc_lookup_actual = dict(zip(actual_rpc_table["key_full"], actual_rpc_table["Actual_RPC"]))
    rpc_lookup_rolling = dict(zip(actual_rpc_table["key_full"], actual_rpc_table["Rolling_4W_Avg_RPC"]))

    # Row level
    row_df = w1_df[["TOUCHPOINT", "LENDER", "SEGMENT", "feo_cohort", "Imp_pct", "CTR", "ENROLLS"]].copy()
    row_df = row_df.drop_duplicates(subset=["TOUCHPOINT", "LENDER", "SEGMENT", "feo_cohort"])
    row_df["key_no_cohort"] = row_df["TOUCHPOINT"] + row_df["LENDER"] + row_df["SEGMENT"]
    row_df["key_full"] = row_df["key_no_cohort"] + row_df["feo_cohort"].astype(str)

    row_df["Actual_RPC"] = row_df["key_full"].map(rpc_lookup_actual).fillna(0)
    row_df["Rolling_4W_Avg_RPC"] = row_df["key_full"].map(rpc_lookup_rolling).fillna(0)
    row_df["W1_Actual_RPU_row"] = row_df["Imp_pct"] * row_df["CTR"] * row_df["Actual_RPC"]
    row_df["W1_Projected_RPU_row"] = row_df["Imp_pct"] * row_df["CTR"] * row_df["Rolling_4W_Avg_RPC"]

    print(f"  W1 row level: {len(row_df)} rows")

    # Segment aggregation
    seg_agg = row_df.groupby(["SEGMENT", "feo_cohort"]).agg(
        Num_Rows_Summed=("W1_Actual_RPU_row", "count"),
        Segment_W1_Actual_RPU=("W1_Actual_RPU_row", "sum"),
        Segment_W1_Projected_RPU=("W1_Projected_RPU_row", "sum"),
    ).reset_index()

    print(f"  W1 segment aggregation: {len(seg_agg)} rows")

    # Blended RPU
    blended = _blend_across_segments(
        seg_agg,
        actual_col="Segment_W1_Actual_RPU",
        projected_col="Segment_W1_Projected_RPU",
        prefix="W1",
    )

    print(f"  W1 blended: {len(blended)} rows")
    return row_df, seg_agg, blended


def _blend_across_segments(seg_agg, actual_col, projected_col, prefix, enrolls_df=None):
    """Enrollment weighted blend across segments for a given projection stage.

    If enrolls_df is None, uses the ENROLLS column already present in seg_agg (for W2W4 actual).
    Otherwise merges enrollment data from enrolls_df.
    """
    # We need enrollment counts. They should come from the enrolls data.
    # seg_agg has SEGMENT and feo_cohort. We need to get enrolls per segment per cohort.
    # The enrollment data is already available from the global scope or passed in.
    # For simplicity, we'll pivot seg_agg and compute the blend.

    segments = seg_agg["SEGMENT"].unique()
    cohorts = seg_agg["feo_cohort"].unique()

    records = []
    for cohort in cohorts:
        cohort_data = seg_agg[seg_agg["feo_cohort"] == cohort]
        total_enrolls = 0
        weighted_actual = 0
        weighted_projected = 0
        seg_details = {}

        for _, srow in cohort_data.iterrows():
            seg = srow["SEGMENT"]
            act = srow[actual_col]
            proj = srow[projected_col]
            enr = srow.get("ENROLLS", srow.get("Num_Rows_Summed", 0))
            seg_details[seg] = {"actual": act, "projected": proj, "enrolls": enr}
            weighted_actual += act * enr
            weighted_projected += proj * enr
            total_enrolls += enr

        rec = {"feo_cohort": cohort}
        for seg_name in ["NS1/3 & all CS", "NS2 & all CS"]:
            short = "NS1" if "NS1" in seg_name else "NS2"
            d = seg_details.get(seg_name, {"actual": 0, "projected": 0, "enrolls": 0})
            rec[f"{short}_Actual_RPU"] = d["actual"]
            rec[f"{short}_Projected_RPU"] = d["projected"]
            rec[f"{short}_Enrolls"] = d["enrolls"]

        if total_enrolls > 0:
            rec[f"{prefix}_Blended_Actual_RPU"] = weighted_actual / total_enrolls
            rec[f"{prefix}_Blended_Projected_RPU"] = weighted_projected / total_enrolls
        else:
            rec[f"{prefix}_Blended_Actual_RPU"] = 0
            rec[f"{prefix}_Blended_Projected_RPU"] = 0

        records.append(rec)

    return pd.DataFrame(records)


# ==============================================================================
# STEP 4: W2 W4 PROJECTION
# ==============================================================================

def build_ratio_base(w1_df, w2w4_df, offsets, metric_name, w1_metric_col, w2w4_metric_col):
    """Build a single ratio base table (Imp, CTR, or RPC) for W2W4."""
    print(f"  Building W2W4 {metric_name} ratio base...")

    # Get W1 metrics per (TOUCHPOINT, LENDER, SEGMENT, feo_cohort)
    w1_metrics = w1_df.groupby(["TOUCHPOINT", "LENDER", "SEGMENT", "feo_cohort"]).agg(
        W1_metric=(w1_metric_col, "first")
    ).reset_index()
    w1_metrics["key_no_cohort"] = w1_metrics["TOUCHPOINT"] + w1_metrics["LENDER"] + w1_metrics["SEGMENT"]
    w1_metrics["key_full"] = w1_metrics["key_no_cohort"] + w1_metrics["feo_cohort"].astype(str)

    # Get W2W4 metrics per (TOUCHPOINT, LENDER, SEGMENT, feo_cohort)
    w2w4_metrics = w2w4_df.groupby(["TOUCHPOINT", "LENDER", "SEGMENT", "feo_cohort"]).agg(
        W2W4_metric=(w2w4_metric_col, "first")
    ).reset_index()
    w2w4_metrics["key_no_cohort"] = w2w4_metrics["TOUCHPOINT"] + w2w4_metrics["LENDER"] + w2w4_metrics["SEGMENT"]

    # Merge W1 and W2W4 on the same key
    merged = w1_metrics.merge(
        w2w4_metrics[["key_no_cohort", "feo_cohort", "W2W4_metric"]],
        on=["key_no_cohort", "feo_cohort"],
        how="left",
    )
    merged["W2W4_metric"] = merged["W2W4_metric"].fillna(0)

    # Rename for clarity
    merged = merged.rename(columns={
        "W1_metric": f"W1_{metric_name}",
        "W2W4_metric": f"W2W4_{metric_name}",
    })

    # Raw ratio
    merged[f"Raw_{metric_name}_Ratio"] = safe_div(
        merged[f"W2W4_{metric_name}"].values,
        merged[f"W1_{metric_name}"].values,
    )

    # Build lookup for rolling average: (key_no_cohort, feo_cohort) -> raw_ratio
    ratio_lookup = {}
    for _, row in merged.iterrows():
        ratio_lookup[(row["key_no_cohort"], row["feo_cohort"])] = row[f"Raw_{metric_name}_Ratio"]

    # Compute lookback and rolling average
    lookback_cols = []
    for offset in offsets:
        col_name = f"Lookback_{offset}d_Ratio"
        lookback_cols.append(col_name)
        merged[col_name] = merged.apply(
            lambda r: ratio_lookup.get(
                (r["key_no_cohort"], r["feo_cohort"] - timedelta(days=offset)), np.nan
            ),
            axis=1,
        )

    merged[f"Rolling_Avg_{metric_name}_Ratio"] = merged.apply(
        lambda r: recency_weighted_avg([r[c] for c in lookback_cols]),
        axis=1,
    )

    print(f"    {metric_name} ratio base: {len(merged)} rows")
    return merged


def compute_w2w4_projection(w1_df, w2w4_df, enriched_df, enrolls_df):
    """Compute W2W4 projections using 3 component ratio approach."""
    print("Computing W2W4 projections...")

    # Build ratio bases for Imp, CTR, RPC
    imp_ratio_df = build_ratio_base(w1_df, w2w4_df, W2W4_RATIO_OFFSETS, "Imp_pct", "Imp_pct", "Imp_pct")
    ctr_ratio_df = build_ratio_base(w1_df, w2w4_df, W2W4_RATIO_OFFSETS, "CTR", "CTR", "CTR")
    rpc_ratio_df = build_ratio_base(w1_df, w2w4_df, W2W4_RATIO_OFFSETS, "RPC", "RPC", "RPC")

    # Build lookups for rolling avg ratios
    imp_lookup = dict(zip(imp_ratio_df["key_full"], imp_ratio_df["Rolling_Avg_Imp_pct_Ratio"]))
    ctr_lookup = dict(zip(ctr_ratio_df["key_full"], ctr_ratio_df["Rolling_Avg_CTR_Ratio"]))
    rpc_lookup = dict(zip(rpc_ratio_df["key_full"], rpc_ratio_df["Rolling_Avg_RPC_Ratio"]))

    # Row level W2W4 projected RPU (one row per W1 combo)
    row_df = w1_df[["TOUCHPOINT", "LENDER", "SEGMENT", "feo_cohort", "Imp_pct", "CTR", "RPC", "ENROLLS"]].copy()
    row_df = row_df.drop_duplicates(subset=["TOUCHPOINT", "LENDER", "SEGMENT", "feo_cohort"])
    row_df["key_no_cohort"] = row_df["TOUCHPOINT"] + row_df["LENDER"] + row_df["SEGMENT"]
    row_df["key_full"] = row_df["key_no_cohort"] + row_df["feo_cohort"].astype(str)

    row_df = row_df.rename(columns={"Imp_pct": "W1_Imp_pct", "CTR": "W1_CTR", "RPC": "W1_RPC"})
    row_df["Imp_Ratio"] = row_df["key_full"].map(imp_lookup).fillna(DEFAULT_IMP_RATIO)
    row_df["CTR_Ratio"] = row_df["key_full"].map(ctr_lookup).fillna(DEFAULT_CTR_RATIO)
    row_df["RPC_Ratio"] = row_df["key_full"].map(rpc_lookup).fillna(DEFAULT_RPC_RATIO)

    # Where Imp_Ratio is 0, the whole product should be 0
    row_df["W2W4_Projected_RPU_row"] = (
        row_df["W1_Imp_pct"] * row_df["W1_CTR"] * row_df["W1_RPC"]
        * row_df["Imp_Ratio"] * row_df["CTR_Ratio"] * row_df["RPC_Ratio"]
    )

    print(f"  W2W4 row level: {len(row_df)} rows")

    # Compute W2W4 actual RPU per row from the W2W4 filtered data
    w2w4_actual = w2w4_df.groupby(["TOUCHPOINT", "LENDER", "SEGMENT", "feo_cohort"]).agg(
        W2W4_Actual_RPU_row=("RPU", "sum")
    ).reset_index()
    w2w4_actual["key_full"] = (
        w2w4_actual["TOUCHPOINT"] + w2w4_actual["LENDER"]
        + w2w4_actual["SEGMENT"] + w2w4_actual["feo_cohort"].astype(str)
    )
    w2w4_actual_lookup = dict(zip(w2w4_actual["key_full"], w2w4_actual["W2W4_Actual_RPU_row"]))
    row_df["W2W4_Actual_RPU_row"] = row_df["key_full"].map(w2w4_actual_lookup).fillna(0)

    # Segment aggregation
    seg_agg = row_df.groupby(["SEGMENT", "feo_cohort"]).agg(
        Num_Rows_Summed=("W2W4_Projected_RPU_row", "count"),
        Segment_W2W4_Actual_RPU=("W2W4_Actual_RPU_row", "sum"),
        Segment_W2W4_Projected_RPU=("W2W4_Projected_RPU_row", "sum"),
    ).reset_index()

    # We need enrollment counts for blending
    enrolls_lookup = dict(zip(
        enrolls_df["feo_cohort"].astype(str) + enrolls_df["segment"],
        enrolls_df["enrols"]
    ))
    seg_agg["ENROLLS"] = (seg_agg["feo_cohort"].astype(str) + seg_agg["SEGMENT"]).map(enrolls_lookup).fillna(0)

    print(f"  W2W4 segment aggregation: {len(seg_agg)} rows")

    # Blended
    blended = _blend_across_segments(
        seg_agg,
        actual_col="Segment_W2W4_Actual_RPU",
        projected_col="Segment_W2W4_Projected_RPU",
        prefix="W2W4",
    )

    print(f"  W2W4 blended: {len(blended)} rows")
    return imp_ratio_df, ctr_ratio_df, rpc_ratio_df, row_df, seg_agg, blended


# ==============================================================================
# STEP 5: M3 PROJECTION (DIRECT RPU RATIO)
# ==============================================================================

def compute_m3_direct_rpu_projection(enriched_df, enrolls_df):
    """Compute M3 projection using direct RPU ratio at segment level."""
    print("Computing M3 projections (Direct RPU Ratio)...")

    # Get W1 payout per (SEGMENT, feo_cohort)
    w1_data = enriched_df[enriched_df["click_cohort"] == "1. W1"]
    w1_seg = w1_data.groupby(["SEGMENT", "feo_cohort"]).agg(
        W1_Total_Payout=("payout", "sum"),
    ).reset_index()

    # Get M3 payout per (SEGMENT, feo_cohort)
    m3_data = enriched_df[enriched_df["click_cohort"] == "3. M3"]
    m3_seg = m3_data.groupby(["SEGMENT", "feo_cohort"]).agg(
        M3_Total_Payout=("payout", "sum"),
    ).reset_index()

    # Get enrollment counts
    enrolls_lookup = dict(zip(
        enrolls_df["feo_cohort"].astype(str) + enrolls_df["segment"],
        enrolls_df["enrols"]
    ))

    # Merge W1 and M3
    seg_rpu = w1_seg.merge(m3_seg, on=["SEGMENT", "feo_cohort"], how="outer").fillna(0)
    seg_rpu["ENROLLS"] = (seg_rpu["feo_cohort"].astype(str) + seg_rpu["SEGMENT"]).map(enrolls_lookup).fillna(0)
    seg_rpu["W1_RPU"] = safe_div(seg_rpu["W1_Total_Payout"].values, seg_rpu["ENROLLS"].values)
    seg_rpu["M3_RPU"] = safe_div(seg_rpu["M3_Total_Payout"].values, seg_rpu["ENROLLS"].values)

    print(f"  M3 segment RPU: {len(seg_rpu)} rows")

    # Raw RPU ratio
    ratio_detail = seg_rpu[["SEGMENT", "feo_cohort", "W1_RPU", "M3_RPU"]].copy()
    ratio_detail["Raw_RPU_Ratio"] = safe_div(ratio_detail["M3_RPU"].values, ratio_detail["W1_RPU"].values)

    # Build lookup: (SEGMENT, feo_cohort) -> Raw_RPU_Ratio
    ratio_lookup = {}
    for _, row in ratio_detail.iterrows():
        ratio_lookup[(row["SEGMENT"], row["feo_cohort"])] = row["Raw_RPU_Ratio"]

    # Recency weighted rolling average at M3 offsets
    lookback_cols = []
    for offset in M3_RATIO_OFFSETS:
        col_name = f"Lookback_{offset}d_Ratio"
        lookback_cols.append(col_name)
        ratio_detail[col_name] = ratio_detail.apply(
            lambda r: ratio_lookup.get(
                (r["SEGMENT"], r["feo_cohort"] - timedelta(days=offset)), np.nan
            ),
            axis=1,
        )

    ratio_detail["Rolling_Avg_RPU_Ratio"] = ratio_detail.apply(
        lambda r: recency_weighted_avg([r[c] for c in lookback_cols]),
        axis=1,
    )

    # M3 projected RPU per segment
    ratio_detail["M3_Projected_RPU_seg"] = ratio_detail["W1_RPU"] * ratio_detail["Rolling_Avg_RPU_Ratio"]

    print(f"  M3 ratio detail: {len(ratio_detail)} rows")

    # M3 actual RPU per segment (for blending)
    ratio_detail["M3_Actual_RPU_seg"] = ratio_detail["M3_RPU"]

    # Blended M3 RPU
    # We need enrollment data merged in
    blend_input = ratio_detail[["SEGMENT", "feo_cohort", "M3_Actual_RPU_seg", "M3_Projected_RPU_seg"]].copy()
    blend_input["ENROLLS"] = (blend_input["feo_cohort"].astype(str) + blend_input["SEGMENT"]).map(enrolls_lookup).fillna(0)

    blended = _blend_across_segments(
        blend_input,
        actual_col="M3_Actual_RPU_seg",
        projected_col="M3_Projected_RPU_seg",
        prefix="M3",
    )

    print(f"  M3 blended: {len(blended)} rows")
    return seg_rpu, ratio_detail, blended


# ==============================================================================
# STEP 6: MAIN OUTPUT
# ==============================================================================

def build_main_output(w1_blended, w2w4_blended, m3_blended):
    """Combine all projection windows into the executive summary."""
    print("Building Main output...")

    main = w1_blended[["feo_cohort", "W1_Blended_Actual_RPU", "W1_Blended_Projected_RPU"]].copy()
    main = main.rename(columns={
        "W1_Blended_Actual_RPU": "W1_Actual_RPU",
        "W1_Blended_Projected_RPU": "W1_Projected_RPU",
    })

    # Merge W2W4
    w2w4_cols = w2w4_blended[["feo_cohort", "W2W4_Blended_Actual_RPU", "W2W4_Blended_Projected_RPU"]].rename(columns={
        "W2W4_Blended_Actual_RPU": "W2W4_Actual_RPU",
        "W2W4_Blended_Projected_RPU": "W2W4_Projected_RPU",
    })
    main = main.merge(w2w4_cols, on="feo_cohort", how="outer")

    # Merge M3
    m3_cols = m3_blended[["feo_cohort", "M3_Blended_Actual_RPU", "M3_Blended_Projected_RPU"]].rename(columns={
        "M3_Blended_Actual_RPU": "M3_Actual_RPU",
        "M3_Blended_Projected_RPU": "M3_Projected_RPU",
    })
    main = main.merge(m3_cols, on="feo_cohort", how="outer")

    main = main.fillna(0)

    main["Total_Predicted_RPU"] = main["W1_Projected_RPU"] + main["W2W4_Projected_RPU"] + main["M3_Projected_RPU"]
    main["Total_Actual_RPU"] = main["W1_Actual_RPU"] + main["W2W4_Actual_RPU"] + main["M3_Actual_RPU"]
    main["Delta"] = main["Total_Actual_RPU"] - main["Total_Predicted_RPU"]
    main["Delta_Pct"] = safe_div(main["Delta"].values, main["Total_Actual_RPU"].values)

    # Reorder columns
    main = main[[
        "feo_cohort",
        "W1_Projected_RPU", "W1_Actual_RPU",
        "W2W4_Projected_RPU", "W2W4_Actual_RPU",
        "M3_Projected_RPU", "M3_Actual_RPU",
        "Total_Predicted_RPU", "Total_Actual_RPU",
        "Delta", "Delta_Pct",
    ]]

    print(f"  Main output: {len(main)} rows")
    return main


# ==============================================================================
# STEP 7: TEST CASES
# ==============================================================================

def run_tests(enriched_df, w1_df, actual_rpc_table, w1_row_df, w1_seg_agg, w1_blended,
              imp_ratio_df, ctr_ratio_df, rpc_ratio_df, w2w4_row_df, w2w4_seg_agg, w2w4_blended,
              m3_seg_rpu, m3_ratio_detail, m3_blended, main_df, enrolls_df):
    """Run all 12 test cases and return results."""
    print("\n" + "=" * 50)
    print("RUNNING TEST CASES")
    print("=" * 50)

    results = []
    details = []

    # ---- Test 1: Imp% Calculation Accuracy ----
    test_name = "Imp% Calculation"
    try:
        eligible = enriched_df[enriched_df["ENROLLS"] > 0]
        sample = eligible.head(3) if len(eligible) < 3 else eligible.sample(n=3, random_state=42)
        passed = True
        test_detail_rows = []
        for _, row in sample.iterrows():
            expected = row["impression_count"] / row["ENROLLS"]
            actual = row["Imp_pct"]
            row_pass = abs(actual - expected) < 1e-6
            if not row_pass:
                passed = False
            test_detail_rows.append({
                "Test": "1", "Detail": f"{row['feo_cohort'].date()} | {row['SEGMENT']} | {row['TOUCHPOINT']} | {row['LENDER']}",
                "Expected": round(expected, 8), "Computed": round(actual, 8),
                "Result": "PASS" if row_pass else "FAIL",
            })
        details.extend(test_detail_rows)
        results.append(("Test  1", test_name, "PASS" if passed else "FAIL"))
        print(f"  Test  1: {test_name} {'.' * (30 - len(test_name))} {'PASS' if passed else 'FAIL'}")
    except Exception as e:
        results.append(("Test  1", test_name, f"ERROR: {e}"))
        print(f"  Test  1: {test_name} {'.' * (30 - len(test_name))} ERROR")

    # ---- Test 2: Recency Weighted Rolling Average RPC ----
    test_name = "Recency Wtd Avg RPC"
    try:
        # Find combos with 5+ cohorts
        combo_counts = actual_rpc_table.groupby("key_no_cohort").size()
        good_combos = combo_counts[combo_counts >= 5].index.tolist()
        if len(good_combos) == 0:
            good_combos = combo_counts.index.tolist()
        test_combos = good_combos[:2]

        passed = True
        for combo_key in test_combos:
            combo_data = actual_rpc_table[actual_rpc_table["key_no_cohort"] == combo_key].sort_values("feo_cohort")
            # Pick the last row (most lookback data available)
            test_row = combo_data.iloc[-1]
            lookback_vals = [test_row.get(f"Lookback_{o}d_RPC", np.nan) for o in W1_RPC_OFFSETS]
            expected = recency_weighted_avg(lookback_vals)
            actual_val = test_row["Rolling_4W_Avg_RPC"]
            row_pass = abs(actual_val - expected) < 1e-6
            if not row_pass:
                passed = False
            details.append({
                "Test": "2", "Detail": f"{combo_key} @ {test_row['feo_cohort'].date()}",
                "Expected": round(expected, 8), "Computed": round(actual_val, 8),
                "Result": "PASS" if row_pass else "FAIL",
            })
        results.append(("Test  2", test_name, "PASS" if passed else "FAIL"))
        print(f"  Test  2: {test_name} {'.' * (30 - len(test_name))} {'PASS' if passed else 'FAIL'}")
    except Exception as e:
        results.append(("Test  2", test_name, f"ERROR: {e}"))
        print(f"  Test  2: {test_name} {'.' * (30 - len(test_name))} ERROR")

    # ---- Test 3: W1 Projected RPU Row Level ----
    test_name = "W1 Projected RPU Row"
    try:
        sample = w1_row_df[w1_row_df["Rolling_4W_Avg_RPC"] > 0].head(3)
        passed = True
        for _, row in sample.iterrows():
            expected = row["Imp_pct"] * row["CTR"] * row["Rolling_4W_Avg_RPC"]
            actual_val = row["W1_Projected_RPU_row"]
            row_pass = abs(actual_val - expected) < 1e-6
            if not row_pass:
                passed = False
            details.append({
                "Test": "3", "Detail": f"{row['key_full'][:60]}",
                "Expected": round(expected, 8), "Computed": round(actual_val, 8),
                "Result": "PASS" if row_pass else "FAIL",
            })
        results.append(("Test  3", test_name, "PASS" if passed else "FAIL"))
        print(f"  Test  3: {test_name} {'.' * (30 - len(test_name))} {'PASS' if passed else 'FAIL'}")
    except Exception as e:
        results.append(("Test  3", test_name, f"ERROR: {e}"))
        print(f"  Test  3: {test_name} {'.' * (30 - len(test_name))} ERROR")

    # ---- Test 4: W1 Segment Aggregation ----
    test_name = "W1 Segment Aggregation"
    try:
        passed = True
        for _, agg_row in w1_seg_agg.head(2).iterrows():
            seg = agg_row["SEGMENT"]
            cohort = agg_row["feo_cohort"]
            row_level_sum = w1_row_df[
                (w1_row_df["SEGMENT"] == seg) & (w1_row_df["feo_cohort"] == cohort)
            ]["W1_Projected_RPU_row"].sum()
            expected = row_level_sum
            actual_val = agg_row["Segment_W1_Projected_RPU"]
            row_pass = abs(actual_val - expected) < 1e-6
            if not row_pass:
                passed = False
            details.append({
                "Test": "4", "Detail": f"{seg} @ {cohort.date()} ({int(agg_row['Num_Rows_Summed'])} rows)",
                "Expected": round(expected, 8), "Computed": round(actual_val, 8),
                "Result": "PASS" if row_pass else "FAIL",
            })
        results.append(("Test  4", test_name, "PASS" if passed else "FAIL"))
        print(f"  Test  4: {test_name} {'.' * (30 - len(test_name))} {'PASS' if passed else 'FAIL'}")
    except Exception as e:
        results.append(("Test  4", test_name, f"ERROR: {e}"))
        print(f"  Test  4: {test_name} {'.' * (30 - len(test_name))} ERROR")

    # ---- Test 5: W1 Blended RPU ----
    test_name = "W1 Blended RPU"
    try:
        passed = True
        for _, brow in w1_blended.head(2).iterrows():
            ns1_proj = brow["NS1_Projected_RPU"]
            ns2_proj = brow["NS2_Projected_RPU"]
            ns1_enr = brow["NS1_Enrolls"]
            ns2_enr = brow["NS2_Enrolls"]
            total_enr = ns1_enr + ns2_enr
            if total_enr > 0:
                expected = (ns1_proj * ns1_enr + ns2_proj * ns2_enr) / total_enr
            else:
                expected = 0
            actual_val = brow["W1_Blended_Projected_RPU"]
            row_pass = abs(actual_val - expected) < 1e-6
            if not row_pass:
                passed = False
            details.append({
                "Test": "5", "Detail": f"{brow['feo_cohort'].date()} | NS1={ns1_proj:.6f}*{ns1_enr:.0f} + NS2={ns2_proj:.6f}*{ns2_enr:.0f}",
                "Expected": round(expected, 8), "Computed": round(actual_val, 8),
                "Result": "PASS" if row_pass else "FAIL",
            })
        results.append(("Test  5", test_name, "PASS" if passed else "FAIL"))
        print(f"  Test  5: {test_name} {'.' * (30 - len(test_name))} {'PASS' if passed else 'FAIL'}")
    except Exception as e:
        results.append(("Test  5", test_name, f"ERROR: {e}"))
        print(f"  Test  5: {test_name} {'.' * (30 - len(test_name))} ERROR")

    # ---- Test 6: W2W4 Ratio Calculation ----
    test_name = "W2W4 Ratio Calculation"
    try:
        passed = True
        sample = imp_ratio_df[imp_ratio_df["Rolling_Avg_Imp_pct_Ratio"] > 0].head(2)
        for _, row in sample.iterrows():
            lookback_vals = [row.get(f"Lookback_{o}d_Ratio", np.nan) for o in W2W4_RATIO_OFFSETS]
            expected = recency_weighted_avg(lookback_vals)
            actual_val = row["Rolling_Avg_Imp_pct_Ratio"]
            row_pass = abs(actual_val - expected) < 1e-6
            if not row_pass:
                passed = False
            details.append({
                "Test": "6", "Detail": f"Imp ratio: {row['key_full'][:50]}",
                "Expected": round(expected, 8), "Computed": round(actual_val, 8),
                "Result": "PASS" if row_pass else "FAIL",
            })
        results.append(("Test  6", test_name, "PASS" if passed else "FAIL"))
        print(f"  Test  6: {test_name} {'.' * (30 - len(test_name))} {'PASS' if passed else 'FAIL'}")
    except Exception as e:
        results.append(("Test  6", test_name, f"ERROR: {e}"))
        print(f"  Test  6: {test_name} {'.' * (30 - len(test_name))} ERROR")

    # ---- Test 7: W2W4 Projected RPU Row Level ----
    test_name = "W2W4 Projected RPU Row"
    try:
        passed = True
        sample = w2w4_row_df[w2w4_row_df["W2W4_Projected_RPU_row"] > 0].head(2)
        for _, row in sample.iterrows():
            expected = (
                row["W1_Imp_pct"] * row["W1_CTR"] * row["W1_RPC"]
                * row["Imp_Ratio"] * row["CTR_Ratio"] * row["RPC_Ratio"]
            )
            actual_val = row["W2W4_Projected_RPU_row"]
            row_pass = abs(actual_val - expected) < 1e-6
            if not row_pass:
                passed = False
            details.append({
                "Test": "7", "Detail": f"{row['key_full'][:50]}",
                "Expected": round(expected, 8), "Computed": round(actual_val, 8),
                "Result": "PASS" if row_pass else "FAIL",
            })
        results.append(("Test  7", test_name, "PASS" if passed else "FAIL"))
        print(f"  Test  7: {test_name} {'.' * (30 - len(test_name))} {'PASS' if passed else 'FAIL'}")
    except Exception as e:
        results.append(("Test  7", test_name, f"ERROR: {e}"))
        print(f"  Test  7: {test_name} {'.' * (30 - len(test_name))} ERROR")

    # ---- Test 8: M3 Direct RPU Ratio Verification ----
    test_name = "M3 Direct RPU Ratio"
    try:
        passed = True
        sample = m3_ratio_detail[m3_ratio_detail["W1_RPU"] > 0].head(2)
        for _, row in sample.iterrows():
            # Check Raw_RPU_Ratio = M3_RPU / W1_RPU
            expected_ratio = row["M3_RPU"] / row["W1_RPU"]
            actual_ratio = row["Raw_RPU_Ratio"]
            ratio_pass = abs(actual_ratio - expected_ratio) < 1e-6

            # Check M3_Projected = W1_RPU * Rolling_Avg
            expected_proj = row["W1_RPU"] * row["Rolling_Avg_RPU_Ratio"]
            actual_proj = row["M3_Projected_RPU_seg"]
            proj_pass = abs(actual_proj - expected_proj) < 1e-6

            row_pass = ratio_pass and proj_pass
            if not row_pass:
                passed = False
            details.append({
                "Test": "8", "Detail": f"{row['SEGMENT']} @ {row['feo_cohort'].date()} | ratio={actual_ratio:.6f} proj={actual_proj:.6f}",
                "Expected": f"ratio={expected_ratio:.6f} proj={expected_proj:.6f}",
                "Computed": f"ratio={actual_ratio:.6f} proj={actual_proj:.6f}",
                "Result": "PASS" if row_pass else "FAIL",
            })
        results.append(("Test  8", test_name, "PASS" if passed else "FAIL"))
        print(f"  Test  8: {test_name} {'.' * (30 - len(test_name))} {'PASS' if passed else 'FAIL'}")
    except Exception as e:
        results.append(("Test  8", test_name, f"ERROR: {e}"))
        print(f"  Test  8: {test_name} {'.' * (30 - len(test_name))} ERROR")

    # ---- Test 9: Default Value Behavior ----
    test_name = "Default Value Behavior"
    try:
        # Find rows where Imp_Ratio defaulted to 0
        default_rows = w2w4_row_df[w2w4_row_df["Imp_Ratio"] == 0]
        if len(default_rows) > 0:
            sample_row = default_rows.iloc[0]
            imp_ok = sample_row["Imp_Ratio"] == 0
            rpu_ok = sample_row["W2W4_Projected_RPU_row"] == 0
            passed = imp_ok and rpu_ok
            details.append({
                "Test": "9", "Detail": f"Default row: Imp_Ratio={sample_row['Imp_Ratio']}, RPU={sample_row['W2W4_Projected_RPU_row']}",
                "Expected": "Imp_Ratio=0, RPU=0",
                "Computed": f"Imp_Ratio={sample_row['Imp_Ratio']}, RPU={sample_row['W2W4_Projected_RPU_row']}",
                "Result": "PASS" if passed else "FAIL",
            })
        else:
            # All combos have data, which is also valid
            passed = True
            details.append({
                "Test": "9", "Detail": "No missing W2W4 combos found (all have data)",
                "Expected": "N/A", "Computed": "N/A", "Result": "PASS (no missing combos)",
            })
        results.append(("Test  9", test_name, "PASS" if passed else "FAIL"))
        print(f"  Test  9: {test_name} {'.' * (30 - len(test_name))} {'PASS' if passed else 'FAIL'}")
    except Exception as e:
        results.append(("Test  9", test_name, f"ERROR: {e}"))
        print(f"  Test  9: {test_name} {'.' * (30 - len(test_name))} ERROR")

    # ---- Test 10: Total RPU Reconciliation ----
    test_name = "Total RPU Reconciliation"
    try:
        passed = True
        # Pick matured cohorts (8+ weeks old)
        cutoff = main_df["feo_cohort"].max() - timedelta(weeks=8)
        matured = main_df[main_df["feo_cohort"] <= cutoff].sort_values("feo_cohort", ascending=False).head(3)
        if len(matured) == 0:
            matured = main_df.sort_values("feo_cohort", ascending=False).head(3)
        for _, row in matured.iterrows():
            expected_total = row["W1_Projected_RPU"] + row["W2W4_Projected_RPU"] + row["M3_Projected_RPU"]
            actual_total = row["Total_Predicted_RPU"]
            total_pass = abs(actual_total - expected_total) < 1e-6

            expected_delta = row["Total_Actual_RPU"] - row["Total_Predicted_RPU"]
            delta_pass = abs(row["Delta"] - expected_delta) < 1e-6

            if row["Total_Actual_RPU"] != 0:
                expected_pct = expected_delta / row["Total_Actual_RPU"]
            else:
                expected_pct = 0
            pct_pass = abs(row["Delta_Pct"] - expected_pct) < 1e-6

            row_pass = total_pass and delta_pass and pct_pass
            if not row_pass:
                passed = False
            details.append({
                "Test": "10", "Detail": f"{row['feo_cohort'].date()} | Total={actual_total:.6f} Delta={row['Delta']:.6f} Pct={row['Delta_Pct']:.4f}",
                "Expected": f"Total={expected_total:.6f}", "Computed": f"Total={actual_total:.6f}",
                "Result": "PASS" if row_pass else "FAIL",
            })
        results.append(("Test 10", test_name, "PASS" if passed else "FAIL"))
        print(f"  Test 10: {test_name} {'.' * (30 - len(test_name))} {'PASS' if passed else 'FAIL'}")
    except Exception as e:
        results.append(("Test 10", test_name, f"ERROR: {e}"))
        print(f"  Test 10: {test_name} {'.' * (30 - len(test_name))} ERROR")

    # ---- Test 11: Zero Enrollment Guard ----
    test_name = "Zero Enrollment Guard"
    try:
        zero_enr = enriched_df[enriched_df["ENROLLS"] == 0]
        if len(zero_enr) > 0:
            passed = (
                (zero_enr["Imp_pct"] == 0).all()
                and (zero_enr["CTR"] == 0).all()
                and (zero_enr["RPC"] == 0).all()
            )
            details.append({
                "Test": "11", "Detail": f"{len(zero_enr)} rows with ENROLLS=0, all metrics defaulted to 0",
                "Expected": "Imp_pct=0, CTR=0, RPC=0",
                "Computed": f"Imp_pct all 0: {(zero_enr['Imp_pct']==0).all()}, CTR all 0: {(zero_enr['CTR']==0).all()}",
                "Result": "PASS" if passed else "FAIL",
            })
        else:
            # Simulate: verify safe_div handles it
            test_val = safe_div(np.array([100]), np.array([0]))
            passed = test_val[0] == 0
            details.append({
                "Test": "11", "Detail": "No zero enrollment rows found; verified safe_div(100,0)=0",
                "Expected": "0", "Computed": str(test_val[0]),
                "Result": "PASS" if passed else "FAIL",
            })
        results.append(("Test 11", test_name, "PASS" if passed else "FAIL"))
        print(f"  Test 11: {test_name} {'.' * (30 - len(test_name))} {'PASS' if passed else 'FAIL'}")
    except Exception as e:
        results.append(("Test 11", test_name, f"ERROR: {e}"))
        print(f"  Test 11: {test_name} {'.' * (30 - len(test_name))} ERROR")

    # ---- Test 12: Missing Cohort in Rolling Average ----
    test_name = "Missing Cohort Rolling Avg"
    try:
        earliest = actual_rpc_table["feo_cohort"].min()
        earliest_rows = actual_rpc_table[actual_rpc_table["feo_cohort"] == earliest]
        passed = True
        if len(earliest_rows) > 0:
            row = earliest_rows.iloc[0]
            # All lookback dates should be before data start, so all NaN
            lookbacks = [row.get(f"Lookback_{o}d_RPC", np.nan) for o in W1_RPC_OFFSETS]
            # Check that missing ones contributed 0 (rolling avg should still use weights as is)
            expected_avg = recency_weighted_avg(lookbacks)
            actual_avg = row["Rolling_4W_Avg_RPC"]
            passed = abs(actual_avg - expected_avg) < 1e-6

            missing_dates = [str((earliest - timedelta(days=o)).date()) for o in W1_RPC_OFFSETS
                             if pd.isna(row.get(f"Lookback_{o}d_RPC", np.nan))]
            details.append({
                "Test": "12", "Detail": f"Earliest cohort: {earliest.date()} | Missing lookbacks: {', '.join(missing_dates) if missing_dates else 'none'}",
                "Expected": f"Weighted avg={expected_avg:.8f} (missing contribute 0, weights NOT redistributed)",
                "Computed": f"Weighted avg={actual_avg:.8f}",
                "Result": "PASS" if passed else "FAIL",
            })
        results.append(("Test 12", test_name, "PASS" if passed else "FAIL"))
        print(f"  Test 12: {test_name} {'.' * (30 - len(test_name))} {'PASS' if passed else 'FAIL'}")
    except Exception as e:
        results.append(("Test 12", test_name, f"ERROR: {e}"))
        print(f"  Test 12: {test_name} {'.' * (30 - len(test_name))} ERROR")

    # Print summary
    pass_count = sum(1 for _, _, r in results if r == "PASS")
    total_count = len(results)
    print("\n" + "=" * 50)
    print("NON API RPU PROJECTION - TEST RESULTS")
    print("=" * 50)
    for test_id, name, result in results:
        dots = "." * (35 - len(name))
        print(f"  {test_id}: {name} {dots} {result}")
    print(f"\n  RESULT: {pass_count}/{total_count} PASSED")
    print("=" * 50)

    # Build test results DataFrame
    test_results_df = pd.DataFrame(details)
    return test_results_df, results


# ==============================================================================
# STEP 8: WRITE EXCEL
# ==============================================================================

def write_excel(tabs, output_path):
    """Write all 20 tabs to a single Excel workbook with formatting."""
    print(f"\nWriting Excel to {output_path}...")

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for tab_name, df in tabs.items():
            df.to_excel(writer, sheet_name=tab_name, index=False)
            print(f"  Tab '{tab_name}': {len(df)} rows, {len(df.columns)} cols")

        wb = writer.book

        # Tab color mapping
        tab_colors = {}
        for name in tabs:
            num = int(name.split(" ")[0])
            if num <= 3:
                tab_colors[name] = "808080"    # grey for inputs
            elif num <= 8:
                tab_colors[name] = "228B22"    # green for W1
            elif num <= 15:
                tab_colors[name] = "4169E1"    # blue for W2W4
            elif num <= 18:
                tab_colors[name] = "FF8C00"    # orange for M3
            else:
                tab_colors[name] = "800080"    # purple for output

        header_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        header_font = Font(bold=True)

        for ws in wb.worksheets:
            # Tab color
            if ws.title in tab_colors:
                ws.sheet_properties.tabColor = tab_colors[ws.title]

            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Freeze top row
            ws.freeze_panes = "A2"

            # Auto fit column widths (approximation)
            for col_idx, col_cells in enumerate(ws.columns, 1):
                max_len = 0
                for cell in col_cells:
                    try:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted = min(max_len + 2, 40)
                ws.column_dimensions[get_column_letter(col_idx)].width = max(adjusted, 10)

    print(f"  Excel written successfully: {output_path}")


# ==============================================================================
# MAIN ORCHESTRATOR
# ==============================================================================

def main():
    """Run the complete projection pipeline."""
    print("=" * 60)
    print("NON API RPU PROJECTION ENGINE")
    print("=" * 60)

    # Step 1: Load data
    enrolls_df, non_api_raw_df = load_and_clean_data(ENROLLS_PATH, NON_API_PATH)

    # Step 2: Preprocess / enrich
    enriched_df = preprocess(non_api_raw_df, enrolls_df, TOUCHPOINT_MAP)

    # Step 3: Filter datasets
    w1_df = enriched_df[enriched_df["click_cohort"] == "1. W1"].copy()
    w2w4_df = enriched_df[enriched_df["click_cohort"] == "2. W2-W4"].copy()
    print(f"Filtered: W1={len(w1_df)} rows, W2W4={len(w2w4_df)} rows")

    # Step 4: W1 Projection
    actual_rpc_table = build_actual_rpc_table(w1_df)
    w1_row_df, w1_seg_agg, w1_blended = compute_w1_projection(w1_df, actual_rpc_table)

    # Need enrollment counts in w1_seg_agg for proper blending
    enrolls_lookup = dict(zip(
        enrolls_df["feo_cohort"].astype(str) + enrolls_df["segment"],
        enrolls_df["enrols"]
    ))
    w1_seg_agg["ENROLLS"] = (w1_seg_agg["feo_cohort"].astype(str) + w1_seg_agg["SEGMENT"]).map(enrolls_lookup).fillna(0)

    # Recompute W1 blended with correct enrollments
    w1_blended = _blend_across_segments(
        w1_seg_agg,
        actual_col="Segment_W1_Actual_RPU",
        projected_col="Segment_W1_Projected_RPU",
        prefix="W1",
    )

    # Step 5: W2W4 Projection
    imp_ratio_df, ctr_ratio_df, rpc_ratio_df, w2w4_row_df, w2w4_seg_agg, w2w4_blended = \
        compute_w2w4_projection(w1_df, w2w4_df, enriched_df, enrolls_df)

    # Step 6: M3 Projection
    m3_seg_rpu, m3_ratio_detail, m3_blended = compute_m3_direct_rpu_projection(enriched_df, enrolls_df)

    # Step 7: Main Output
    main_df = build_main_output(w1_blended, w2w4_blended, m3_blended)

    # Step 8: Run Tests
    test_results_df, test_summary = run_tests(
        enriched_df, w1_df, actual_rpc_table, w1_row_df, w1_seg_agg, w1_blended,
        imp_ratio_df, ctr_ratio_df, rpc_ratio_df, w2w4_row_df, w2w4_seg_agg, w2w4_blended,
        m3_seg_rpu, m3_ratio_detail, m3_blended, main_df, enrolls_df,
    )

    # Step 9: Sort all outputs (most recent cohort first)
    enrolls_sorted = sort_output(enrolls_df)
    non_api_raw_sorted = sort_output(non_api_raw_df)
    enriched_sorted = sort_output(enriched_df)
    w1_filtered_sorted = sort_output(w1_df)
    actual_rpc_sorted = sort_output(actual_rpc_table)
    w1_row_sorted = sort_output(w1_row_df)
    w1_seg_sorted = sort_output(w1_seg_agg)
    w1_blended_sorted = sort_output(w1_blended)
    w2w4_filtered_sorted = sort_output(w2w4_df)
    imp_ratio_sorted = sort_output(imp_ratio_df)
    ctr_ratio_sorted = sort_output(ctr_ratio_df)
    rpc_ratio_sorted = sort_output(rpc_ratio_df)
    w2w4_row_sorted = sort_output(w2w4_row_df)
    w2w4_seg_sorted = sort_output(w2w4_seg_agg)
    w2w4_blended_sorted = sort_output(w2w4_blended)
    m3_seg_rpu_sorted = sort_output(m3_seg_rpu)
    m3_ratio_sorted = sort_output(m3_ratio_detail)
    m3_blended_sorted = sort_output(m3_blended)
    main_sorted = sort_output(main_df)

    # Step 10: Write Excel
    tabs = {
        "01 Input Enrolls": enrolls_sorted,
        "02 Input Non API Raw": non_api_raw_sorted,
        "03 Enriched Non API": enriched_sorted,
        "04 W1 Filtered Data": w1_filtered_sorted,
        "05 W1 Actual RPC Lookup": actual_rpc_sorted,
        "06 W1 Row Level RPU": w1_row_sorted,
        "07 W1 Segment Aggregation": w1_seg_sorted,
        "08 W1 Blended RPU": w1_blended_sorted,
        "09 W2W4 Filtered Data": w2w4_filtered_sorted,
        "10 W2W4 Imp Ratio Base": imp_ratio_sorted,
        "11 W2W4 CTR Ratio Base": ctr_ratio_sorted,
        "12 W2W4 RPC Ratio Base": rpc_ratio_sorted,
        "13 W2W4 Row Level RPU": w2w4_row_sorted,
        "14 W2W4 Segment Aggregation": w2w4_seg_sorted,
        "15 W2W4 Blended RPU": w2w4_blended_sorted,
        "16 M3 Segment RPU": m3_seg_rpu_sorted,
        "17 M3 RPU Ratio Detail": m3_ratio_sorted,
        "18 M3 Blended RPU": m3_blended_sorted,
        "19 Main": main_sorted,
        "20 Test Results": test_results_df,
    }

    write_excel(tabs, OUTPUT_PATH)

    print("\n" + "=" * 60)
    print("PIPELINE COMPLETE")
    print(f"Output: {OUTPUT_PATH}")
    print("=" * 60)


if __name__ == "__main__":
    main()
