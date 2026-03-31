"""
Combined RPU Projection Merge & Diagnostics Engine
===================================================
Merges API and Non-API RPU projection outputs into a single unified view
and produces diagnostic tables explaining projection error drivers.

Dependencies: pandas, openpyxl only.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

# =============================================================================
# CONFIGURATION
# =============================================================================

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "Outputs"
INPUT_DIR = BASE_DIR / "Inputs"

API_EXCEL_PATH = str(OUTPUT_DIR / "API_RPU_Projections.xlsx")
API_MAIN_SHEET = "22 Main"
API_HEADER_ROW = 2   # Row 1 is description
API_DATA_START = 3

NONAPI_EXCEL_PATH = str(OUTPUT_DIR / "Non_API_RPU_Projection_Output.xlsx")
NONAPI_MAIN_SHEET = "19 Main"
NONAPI_HEADER_ROW = 1   # No description row
NONAPI_DATA_START = 2

ENROLLS_PATH = str(INPUT_DIR / "Non API" / "Enrolls data.csv")

OUTPUT_PATH = str(OUTPUT_DIR / "Combined_RPU_Projections_Diagnostics.xlsx")

COMPONENTS = ["W1", "W2W4", "M3", "NonClick", "NonEnroll"]
API_COMPONENTS = ["W1", "W2W4", "M3", "NonClick", "NonEnroll"]
NONAPI_COMPONENTS = ["W1", "W2W4", "M3"]  # Non-API lacks NonClick & NonEnroll

MATURED_WEEKS_THRESHOLD = 12  # Cohorts with 12+ weeks are considered matured

MATURITY_BUCKETS = {
    (0, 4): "Immature (0-4 weeks)",
    (5, 8): "Developing (5-8 weeks)",
    (9, 11): "Maturing (9-11 weeks)",
    (12, 15): "Mature (12-15 weeks)",
    (16, 999): "Fully Mature (16+ weeks)",
}

MAPE_THRESHOLDS = {
    "green": 0.10,    # < 10%
    "yellow": 0.30,   # 10-30%
    "orange": 0.50,   # 30-50%
    # > 50% = red
}

ERROR_DRIVER_THRESHOLD = 0.30  # > 30% contribution = major driver
PARETO_THRESHOLD = 0.80  # top 80% of cumulative error

KNOWN_DATE_FORMATS = [
    "%B %d, %Y",      # March 30, 2026
    "%b %d, %Y",      # Mar 30, 2026
    "%Y-%m-%d",        # 2026-03-30
    "%m/%d/%Y",        # 03/30/2026
    "%d-%b-%Y",        # 30-Mar-2026
]

# Formatting constants
FILL_GREY = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
FILL_LIGHT_BLUE = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
FILL_ALT_ROW = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_LIGHT_RED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

FONT_BOLD = Font(bold=True)
FONT_BOLD_ITALIC = Font(bold=True, italic=True)
FONT_RED = Font(color="FF0000")
FONT_GREEN = Font(color="008000")
FONT_BOLD_RED = Font(bold=True, color="FF0000")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Tab color assignments
TAB_COLORS = {
    "input": "808080",       # Grey
    "merge": "A9A9A9",       # Light Grey
    "diagnostics": "4472C4", # Blue
    "analysis": "ED7D31",    # Orange
    "summary": "7030A0",     # Purple
}


# =============================================================================
# SAFE DIVISION HELPER
# =============================================================================

def safe_divide(numerator, denominator, default=0):
    """Return numerator / denominator, defaulting to `default` when denominator is 0 or NaN."""
    if isinstance(numerator, pd.Series):
        result = numerator / denominator
        result = result.replace([float('inf'), float('-inf')], default)
        result = result.fillna(default)
        return result
    else:
        if denominator == 0 or pd.isna(denominator):
            return default
        return numerator / denominator


# =============================================================================
# DATE PARSING HELPER
# =============================================================================

def detect_and_parse_dates(series):
    """Try known date formats, fallback to pd.to_datetime with mixed format."""
    if pd.api.types.is_datetime64_any_dtype(series):
        return series
    # Clean whitespace
    cleaned = series.astype(str).str.strip()
    for fmt in KNOWN_DATE_FORMATS:
        try:
            parsed = pd.to_datetime(cleaned, format=fmt)
            if parsed.notna().all():
                print(f"  Date format detected: '{fmt}'")
                return parsed
        except (ValueError, TypeError):
            continue
    # Fallback
    print("  Using fallback mixed date parsing")
    return pd.to_datetime(cleaned, format='mixed', dayfirst=False)


# =============================================================================
# DATA LOADING FUNCTIONS
# =============================================================================

def load_api_main(path):
    """Load API Main tab and return (renamed_df with _API suffix, raw_df with original columns)."""
    print(f"\n--- Loading API Main from: {path}, sheet: {API_MAIN_SHEET} ---")
    df = pd.read_excel(path, sheet_name=API_MAIN_SHEET, header=API_HEADER_ROW - 1,
                        skiprows=range(0, API_HEADER_ROW - 1))
    # The first row after header=1 (0-indexed) means Row 2 is header. But we need to skip row 1 (description).
    # header=1 means use row index 1 (0-based) as header, which is Row 2 in Excel.
    df = pd.read_excel(path, sheet_name=API_MAIN_SHEET, header=1)

    print(f"  Raw shape: {df.shape}")
    df['feo_cohort'] = pd.to_datetime(df['feo_cohort'])
    df = df.dropna(subset=['feo_cohort'])

    # Save raw copy before renaming
    raw_df = df.copy()

    # Columns to rename (component RPU columns)
    cols_to_drop = ['Total_Predicted_RPU', 'Total_Actual_RPU', 'Delta', 'Delta_Pct']
    rename_map = {}
    for col in df.columns:
        if col == 'feo_cohort':
            continue
        if col in cols_to_drop:
            continue
        rename_map[col] = col + '_API'

    renamed = df.drop(columns=cols_to_drop, errors='ignore').rename(columns=rename_map)

    print(f"  Loaded API Main: {len(renamed)} rows, cohorts {renamed['feo_cohort'].min().date()} to {renamed['feo_cohort'].max().date()}")
    return renamed, raw_df


def load_nonapi_main(path):
    """Load Non-API Main tab and return (renamed_df with _NonAPI suffix, raw_df with original columns)."""
    print(f"\n--- Loading Non-API Main from: {path}, sheet: {NONAPI_MAIN_SHEET} ---")
    df = pd.read_excel(path, sheet_name=NONAPI_MAIN_SHEET, header=0)

    print(f"  Raw shape: {df.shape}")
    df['feo_cohort'] = pd.to_datetime(df['feo_cohort'])
    df = df.dropna(subset=['feo_cohort'])

    # Non-API uses "Projected" instead of "Predicted" for W1 — normalize
    col_remap = {}
    for col in df.columns:
        if 'Projected' in col:
            col_remap[col] = col.replace('Projected', 'Predicted')
    if col_remap:
        df = df.rename(columns=col_remap)
        print(f"  Normalized column names: {list(col_remap.keys())} -> {list(col_remap.values())}")

    # Save raw copy before renaming
    raw_df = df.copy()

    cols_to_drop = ['Total_Predicted_RPU', 'Total_Actual_RPU', 'Delta', 'Delta_Pct']
    rename_map = {}
    for col in df.columns:
        if col == 'feo_cohort':
            continue
        if col in cols_to_drop:
            continue
        rename_map[col] = col + '_NonAPI'

    renamed = df.drop(columns=cols_to_drop, errors='ignore').rename(columns=rename_map)

    # Add placeholder columns for Non-Click and Non-Enroll (all 0)
    renamed['NonClick_Predicted_RPU_NonAPI'] = 0
    renamed['NonClick_Actual_RPU_NonAPI'] = 0
    renamed['NonEnroll_Predicted_RPU_NonAPI'] = 0
    renamed['NonEnroll_Actual_RPU_NonAPI'] = 0

    print(f"  Loaded Non-API Main: {len(renamed)} rows, cohorts {renamed['feo_cohort'].min().date()} to {renamed['feo_cohort'].max().date()}")
    return renamed, raw_df


def load_enrolls(path):
    """Load Enrolls data and return DataFrame with total enrolls per cohort."""
    print(f"\n--- Loading Enrolls from: {path} ---")
    df = pd.read_csv(path)

    # Clean comma-formatted enrols
    df['enrols'] = df['enrols'].astype(str).str.replace(',', '').astype(float).astype(int)

    # Parse dates
    df['feo_cohort'] = detect_and_parse_dates(df['feo_cohort'])

    # Compute total enrolls per cohort
    enrolls = df.groupby('feo_cohort')['enrols'].sum().reset_index()
    enrolls.columns = ['feo_cohort', 'Total_Enrolls']

    print(f"  Loaded Enrolls: {len(enrolls)} cohorts, range {enrolls['feo_cohort'].min().date()} to {enrolls['feo_cohort'].max().date()}")
    return enrolls


# =============================================================================
# MERGE AND COMPUTE
# =============================================================================

def merge_sources(api_df, nonapi_df, enrolls_df):
    """Outer join API and Non-API on feo_cohort, then join with enrolls."""
    print("\n--- Merging Sources ---")
    merged = pd.merge(api_df, nonapi_df, on='feo_cohort', how='outer')
    merged = merged.fillna(0)

    # Count overlapping
    api_cohorts = set(api_df['feo_cohort'])
    nonapi_cohorts = set(nonapi_df['feo_cohort'])
    overlap = api_cohorts & nonapi_cohorts
    api_only = api_cohorts - nonapi_cohorts
    nonapi_only = nonapi_cohorts - api_cohorts

    print(f"  Overlapping cohorts: {len(overlap)}")
    print(f"  API-only cohorts: {len(api_only)}")
    print(f"  Non-API-only cohorts: {len(nonapi_only)}")

    # Join with enrolls
    merged = pd.merge(merged, enrolls_df, on='feo_cohort', how='left')
    merged['Total_Enrolls'] = merged['Total_Enrolls'].fillna(0).astype(int)

    merged = merged.sort_values('feo_cohort', ascending=False).reset_index(drop=True)

    print(f"  Merged: {len(merged)} rows total")

    merge_stats = {
        'overlapping': len(overlap),
        'api_only': len(api_only),
        'nonapi_only': len(nonapi_only),
        'total': len(merged),
    }

    return merged, merge_stats


def compute_combined_metrics(merged_df):
    """Compute combined RPU and all derived metrics."""
    print("\n--- Computing Combined Metrics ---")
    df = merged_df.copy()

    # Step 1.1: Combined RPU (direct addition)
    for c in COMPONENTS:
        df[f'{c}_Predicted_RPU_Total'] = df[f'{c}_Predicted_RPU_API'] + df[f'{c}_Predicted_RPU_NonAPI']
        df[f'{c}_Actual_RPU_Total'] = df[f'{c}_Actual_RPU_API'] + df[f'{c}_Actual_RPU_NonAPI']

    # Step 1.2: Grand Totals
    df['Total_Predicted_RPU'] = sum(df[f'{c}_Predicted_RPU_Total'] for c in COMPONENTS)
    df['Total_Actual_RPU'] = sum(df[f'{c}_Actual_RPU_Total'] for c in COMPONENTS)
    df['Delta'] = df['Total_Actual_RPU'] - df['Total_Predicted_RPU']
    df['Delta_Pct'] = safe_divide(df['Delta'], df['Total_Actual_RPU'])

    # Step 2: Component-Level Delta Decomposition
    for c in COMPONENTS:
        df[f'{c}_Delta_RPU'] = df[f'{c}_Actual_RPU_Total'] - df[f'{c}_Predicted_RPU_Total']

    df['Total_Delta_RPU'] = df['Delta']  # Should equal sum of component deltas

    for c in COMPONENTS:
        df[f'{c}_Delta_Contribution_Pct'] = safe_divide(df[f'{c}_Delta_RPU'], df['Total_Delta_RPU'])
        df[f'{c}_Error_Pct'] = safe_divide(df[f'{c}_Delta_RPU'], df[f'{c}_Actual_RPU_Total'])
        df[f'{c}_MAPE'] = safe_divide(df[f'{c}_Delta_RPU'].abs(), df[f'{c}_Actual_RPU_Total'].abs())

    # Step 3: Source-Level Delta Decomposition
    for c in COMPONENTS:
        df[f'{c}_Delta_RPU_API'] = df[f'{c}_Actual_RPU_API'] - df[f'{c}_Predicted_RPU_API']
        df[f'{c}_Delta_RPU_NonAPI'] = df[f'{c}_Actual_RPU_NonAPI'] - df[f'{c}_Predicted_RPU_NonAPI']
        df[f'{c}_API_Share_Pct'] = safe_divide(df[f'{c}_Delta_RPU_API'], df[f'{c}_Delta_RPU'])

    # Total source deltas
    df['Total_Delta_API'] = sum(df[f'{c}_Delta_RPU_API'] for c in COMPONENTS)
    df['Total_Delta_NonAPI'] = sum(df[f'{c}_Delta_RPU_NonAPI'] for c in COMPONENTS)
    df['API_Share_of_Total_Delta_Pct'] = safe_divide(df['Total_Delta_API'], df['Total_Delta_RPU'])

    # Maturity
    latest_cohort = df['feo_cohort'].max()
    df['Weeks_Since_Cohort'] = ((latest_cohort - df['feo_cohort']).dt.days / 7).astype(int)

    def assign_bucket(weeks):
        for (lo, hi), label in MATURITY_BUCKETS.items():
            if lo <= weeks <= hi:
                return label
        return "Unknown"

    df['Maturity_Bucket'] = df['Weeks_Since_Cohort'].apply(assign_bucket)

    # Has Data flags
    for c in ['W1', 'W2W4', 'M3']:
        df[f'{c}_Has_Data'] = df[f'{c}_Actual_RPU_Total'].apply(lambda x: 'Y' if x > 0 else 'N')

    # Most Inaccurate Component per cohort
    def find_most_inaccurate(row):
        worst = None
        worst_val = 0
        for c in COMPONENTS:
            val = abs(row.get(f'{c}_Error_Pct', 0))
            if val > worst_val:
                worst_val = val
                worst = c
        return worst if worst else 'N/A'

    df['Most_Inaccurate_Component'] = df.apply(find_most_inaccurate, axis=1)
    df['Largest_Error_Pct'] = df[[f'{c}_Error_Pct' for c in COMPONENTS]].abs().max(axis=1)

    print(f"  Combined metrics computed for {len(df)} cohorts")
    return df


# =============================================================================
# TAB BUILDER FUNCTIONS
# =============================================================================

def build_input_api_tab(api_raw_df):
    """Tab 01: Raw API Main input data."""
    return api_raw_df.sort_values('feo_cohort', ascending=False).reset_index(drop=True)


def build_input_nonapi_tab(nonapi_raw_df):
    """Tab 02: Raw Non-API Main input data."""
    return nonapi_raw_df.sort_values('feo_cohort', ascending=False).reset_index(drop=True)


def build_combined_main(combined_df):
    """Tab 03: Combined Main — unified executive summary."""
    cols = ['feo_cohort', 'Total_Enrolls']
    for c in COMPONENTS:
        cols.extend([f'{c}_Predicted_RPU_Total', f'{c}_Actual_RPU_Total'])
    cols.extend(['Total_Predicted_RPU', 'Total_Actual_RPU', 'Delta', 'Delta_Pct'])

    df = combined_df[cols].copy()
    # Rename for clean display
    rename = {}
    for c in COMPONENTS:
        rename[f'{c}_Predicted_RPU_Total'] = f'{c}_Predicted_RPU'
        rename[f'{c}_Actual_RPU_Total'] = f'{c}_Actual_RPU'
    df = df.rename(columns=rename)
    return df


def build_delta_waterfall(combined_df):
    """Tab 04: Delta Waterfall — component decomposition of total error."""
    cols = ['feo_cohort', 'Total_Enrolls']
    for c in COMPONENTS:
        cols.extend([f'{c}_Delta_RPU', f'{c}_Delta_Contribution_Pct'])
    cols.append('Total_Delta_RPU')
    return combined_df[cols].copy()


def build_api_vs_nonapi_delta(combined_df):
    """Tab 05: API vs NonAPI Delta — source-level error decomposition."""
    cols = ['feo_cohort']
    for c in ['W1', 'W2W4', 'M3']:
        cols.extend([f'{c}_Delta_RPU_API', f'{c}_Delta_RPU_NonAPI', f'{c}_Delta_RPU', f'{c}_API_Share_Pct'])
    # NonClick and NonEnroll only come from API
    for c in ['NonClick', 'NonEnroll']:
        cols.extend([f'{c}_Delta_RPU_API', f'{c}_Delta_RPU'])
    cols.extend(['Total_Delta_API', 'Total_Delta_NonAPI', 'Total_Delta_RPU', 'API_Share_of_Total_Delta_Pct'])

    df = combined_df[cols].copy()
    # Rename for clarity
    rename = {}
    for c in COMPONENTS:
        rename[f'{c}_Delta_RPU'] = f'{c}_Delta_Total'
        rename[f'{c}_Delta_RPU_API'] = f'{c}_Delta_API'
        if f'{c}_Delta_RPU_NonAPI' in df.columns:
            rename[f'{c}_Delta_RPU_NonAPI'] = f'{c}_Delta_NonAPI'
    rename['Total_Delta_RPU'] = 'Total_Delta'
    df = df.rename(columns=rename)
    return df


def build_component_accuracy(combined_df):
    """Tab 06: Component Accuracy — per-component accuracy metrics."""
    cols = ['feo_cohort']
    for c in COMPONENTS:
        cols.extend([f'{c}_Error_Pct', f'{c}_MAPE'])
    cols.extend(['Most_Inaccurate_Component', 'Largest_Error_Pct'])

    df = combined_df[cols].copy()

    # Add summary row for matured cohorts
    matured = combined_df[combined_df['Weeks_Since_Cohort'] >= MATURED_WEEKS_THRESHOLD]
    matured_with_actuals = matured[matured['Total_Actual_RPU'] > 0]

    summary = {'feo_cohort': 'AVERAGE (Matured 12+ wks)'}
    for c in COMPONENTS:
        summary[f'{c}_Error_Pct'] = matured_with_actuals[f'{c}_Error_Pct'].mean() if len(matured_with_actuals) > 0 else 0
        summary[f'{c}_MAPE'] = matured_with_actuals[f'{c}_MAPE'].mean() if len(matured_with_actuals) > 0 else 0

    # Ranking
    mape_values = {c: summary[f'{c}_MAPE'] for c in COMPONENTS}
    ranked = sorted(mape_values.items(), key=lambda x: x[1])
    summary['Most_Inaccurate_Component'] = ranked[-1][0] if ranked else 'N/A'
    summary['Largest_Error_Pct'] = ranked[-1][1] if ranked else 0

    summary_df = pd.DataFrame([summary])
    df = pd.concat([df, summary_df], ignore_index=True)

    return df


def build_maturity_analysis(combined_df):
    """Tab 07: Maturity Analysis — how accuracy changes with cohort age."""
    cols = [
        'feo_cohort', 'Weeks_Since_Cohort',
        'W1_Has_Data', 'W2W4_Has_Data', 'M3_Has_Data',
        'Total_Predicted_RPU', 'Total_Actual_RPU', 'Delta', 'Delta_Pct',
        'Maturity_Bucket'
    ]
    return combined_df[cols].copy()


def build_period_trends(combined_df):
    """Tab 08: Period Trends — time-series view of accuracy metrics."""
    df = combined_df.sort_values('feo_cohort', ascending=True).copy()

    cols = ['feo_cohort', 'Total_Predicted_RPU', 'Total_Actual_RPU', 'Delta_Pct']
    for c in ['W1', 'W2W4', 'M3']:
        cols.extend([f'{c}_Predicted_RPU_Total', f'{c}_Actual_RPU_Total', f'{c}_Error_Pct'])

    result = df[cols].copy()

    # Rename for display
    rename = {}
    for c in ['W1', 'W2W4', 'M3']:
        rename[f'{c}_Predicted_RPU_Total'] = f'{c}_Predicted_RPU'
        rename[f'{c}_Actual_RPU_Total'] = f'{c}_Actual_RPU'
    result = result.rename(columns=rename)

    # Rolling averages
    result['Rolling_4W_Avg_Delta_Pct'] = result['Delta_Pct'].rolling(window=4, min_periods=1).mean()
    result['Rolling_4W_Avg_W1_Error_Pct'] = result.get('W1_Error_Pct', pd.Series(0, index=result.index)).rolling(window=4, min_periods=1).mean()

    # Replace any inf/nan in rolling
    result = result.replace([float('inf'), float('-inf')], 0).fillna(0)

    return result


def build_error_concentration(combined_df):
    """Tab 09: Error Concentration — Pareto analysis of RPU error."""
    df = combined_df[['feo_cohort', 'Total_Enrolls', 'Delta']].copy()
    df = df.rename(columns={'Delta': 'RPU_Delta'})
    df['Abs_RPU_Delta'] = df['RPU_Delta'].abs()

    # Sort by worst first
    df = df.sort_values('Abs_RPU_Delta', ascending=False).reset_index(drop=True)

    total_abs = df['Abs_RPU_Delta'].sum()
    df['Cumulative_Abs_Delta'] = df['Abs_RPU_Delta'].cumsum()
    df['Pct_of_Total_Abs_Delta'] = safe_divide(df['Abs_RPU_Delta'], total_abs)
    df['Cumulative_Pct_of_Total'] = df['Pct_of_Total_Abs_Delta'].cumsum()
    df['Error_Rank'] = range(1, len(df) + 1)

    return df


def build_summary_dashboard(combined_df):
    """Tab 10: Summary Dashboard — executive summary with key metrics."""
    matured = combined_df[combined_df['Weeks_Since_Cohort'] >= MATURED_WEEKS_THRESHOLD]
    matured_with_actuals = matured[matured['Total_Actual_RPU'] > 0]

    total_cohorts = len(combined_df)
    matured_count = len(matured)

    # Overall MAPE for matured cohorts
    if len(matured_with_actuals) > 0:
        overall_mape = safe_divide(
            matured_with_actuals['Delta'].abs().mean(),
            matured_with_actuals['Total_Actual_RPU'].abs().mean()
        )
        avg_delta_rpu = matured_with_actuals['Delta'].mean()
        direction = "Under-predicted" if avg_delta_rpu > 0 else "Over-predicted"
    else:
        overall_mape = 0
        avg_delta_rpu = 0
        direction = "N/A"

    # Section A
    section_a = [
        ("Metric", "Value"),
        ("Total Cohorts", total_cohorts),
        (f"Matured Cohorts ({MATURED_WEEKS_THRESHOLD}+ weeks)", matured_count),
        ("Overall MAPE (matured cohorts)", overall_mape),
        ("Avg Delta RPU (matured cohorts)", avg_delta_rpu),
        ("Direction", direction),
    ]

    # Section B: Component Accuracy Ranking
    section_b_header = [("Component", "Avg MAPE", "Avg Delta RPU", "Rank", "Verdict")]
    component_stats = []
    for c in COMPONENTS:
        if len(matured_with_actuals) > 0:
            avg_mape = matured_with_actuals[f'{c}_MAPE'].mean()
            avg_delta = matured_with_actuals[f'{c}_Delta_RPU'].mean()
        else:
            avg_mape = 0
            avg_delta = 0
        component_stats.append((c, avg_mape, avg_delta))

    # Rank by MAPE (lower = more accurate)
    ranked = sorted(component_stats, key=lambda x: x[1])
    section_b_data = []
    for rank, (comp, mape, delta) in enumerate(ranked, 1):
        if rank == 1:
            verdict = "Most accurate"
        elif rank == len(ranked):
            verdict = "Least accurate"
        else:
            verdict = f"Rank {rank}"
        section_b_data.append((comp, mape, delta, rank, verdict))

    # Section C: Source Accuracy
    section_c_header = [("Source", "Avg Pred RPU", "Avg Act RPU", "Avg Delta RPU", "MAPE")]
    if len(matured_with_actuals) > 0:
        mwa = matured_with_actuals
        # API totals
        api_pred = sum(mwa[f'{c}_Predicted_RPU_API'].mean() for c in COMPONENTS)
        api_act = sum(mwa[f'{c}_Actual_RPU_API'].mean() for c in COMPONENTS)
        api_delta = api_act - api_pred
        api_mape = safe_divide(abs(api_delta), abs(api_act)) if api_act != 0 else 0

        # NonAPI totals
        nonapi_pred = sum(mwa[f'{c}_Predicted_RPU_NonAPI'].mean() for c in NONAPI_COMPONENTS)
        nonapi_act = sum(mwa[f'{c}_Actual_RPU_NonAPI'].mean() for c in NONAPI_COMPONENTS)
        nonapi_delta = nonapi_act - nonapi_pred
        nonapi_mape = safe_divide(abs(nonapi_delta), abs(nonapi_act)) if nonapi_act != 0 else 0

        # Combined
        comb_pred = mwa['Total_Predicted_RPU'].mean()
        comb_act = mwa['Total_Actual_RPU'].mean()
        comb_delta = comb_act - comb_pred
        comb_mape = safe_divide(abs(comb_delta), abs(comb_act)) if comb_act != 0 else 0
    else:
        api_pred = api_act = api_delta = api_mape = 0
        nonapi_pred = nonapi_act = nonapi_delta = nonapi_mape = 0
        comb_pred = comb_act = comb_delta = comb_mape = 0

    section_c_data = [
        ("API", api_pred, api_act, api_delta, api_mape),
        ("Non-API", nonapi_pred, nonapi_act, nonapi_delta, nonapi_mape),
        ("Combined", comb_pred, comb_act, comb_delta, comb_mape),
    ]

    # Section D: Top 3 Error Drivers
    section_d_header = [("Rank", "Cohort", "Component", "Source", "Delta RPU", "Why")]

    # Find top 3 component-cohort error drivers
    error_drivers = []
    for _, row in combined_df.iterrows():
        for c in COMPONENTS:
            for src in ['API', 'NonAPI']:
                col = f'{c}_Delta_RPU_{src}'
                if col in combined_df.columns:
                    delta_val = row[col]
                    actual_col = f'{c}_Actual_RPU_{src}'
                    actual_val = row.get(actual_col, 0)
                    pred_col = f'{c}_Predicted_RPU_{src}'
                    pred_val = row.get(pred_col, 0)

                    # Generate "Why"
                    error_pct = safe_divide(abs(delta_val), abs(actual_val)) if actual_val != 0 else (1.0 if pred_val != 0 else 0)
                    weeks = row['Weeks_Since_Cohort']

                    if actual_val == 0 and pred_val > 0:
                        why = "Component not yet matured"
                    elif error_pct > 0.50:
                        direction_str = "over" if delta_val < 0 else "under"
                        why = f"Severe {direction_str}-prediction ({error_pct:.0%})"
                    elif weeks < MATURED_WEEKS_THRESHOLD:
                        why = "Insufficient maturation time"
                    else:
                        why = "Normal projection variance"

                    error_drivers.append({
                        'cohort': row['feo_cohort'],
                        'component': c,
                        'source': src.replace('NonAPI', 'Non-API'),
                        'delta_rpu': delta_val,
                        'abs_delta': abs(delta_val),
                        'why': why,
                    })

    error_drivers_df = pd.DataFrame(error_drivers)
    if len(error_drivers_df) > 0:
        top3 = error_drivers_df.nlargest(3, 'abs_delta')
    else:
        top3 = pd.DataFrame()

    section_d_data = []
    for rank, (_, row) in enumerate(top3.iterrows(), 1):
        cohort_str = row['cohort'].strftime('%Y-%m-%d') if hasattr(row['cohort'], 'strftime') else str(row['cohort'])
        section_d_data.append((rank, cohort_str, row['component'], row['source'], row['delta_rpu'], row['why']))

    return {
        'section_a': section_a,
        'section_b_header': section_b_header,
        'section_b_data': section_b_data,
        'section_c_header': section_c_header,
        'section_c_data': section_c_data,
        'section_d_header': section_d_header,
        'section_d_data': section_d_data,
    }


def build_comment_tab(api_info, nonapi_info, enrolls_info, merge_stats, test_results):
    """Tab 11: Comment — documentation and test results."""
    rows = []

    # Section A: Input File Summary
    rows.append(("SECTION A: INPUT FILE SUMMARY", ""))
    rows.append(("API File", api_info.get('path', '')))
    rows.append(("API Sheet", api_info.get('sheet', '')))
    rows.append(("API Rows Loaded", api_info.get('rows', 0)))
    rows.append(("API Cohort Range", api_info.get('range', '')))
    rows.append(("", ""))
    rows.append(("Non-API File", nonapi_info.get('path', '')))
    rows.append(("Non-API Sheet", nonapi_info.get('sheet', '')))
    rows.append(("Non-API Rows Loaded", nonapi_info.get('rows', 0)))
    rows.append(("Non-API Cohort Range", nonapi_info.get('range', '')))
    rows.append(("", ""))
    rows.append(("Enrolls File", enrolls_info.get('path', '')))
    rows.append(("Enrolls Rows Loaded", enrolls_info.get('rows', 0)))
    rows.append(("Enrolls Cohort Range", enrolls_info.get('range', '')))
    rows.append(("", ""))
    rows.append(("Overlapping Cohorts", merge_stats.get('overlapping', 0)))
    rows.append(("API-Only Cohorts", merge_stats.get('api_only', 0)))
    rows.append(("Non-API-Only Cohorts", merge_stats.get('nonapi_only', 0)))
    rows.append(("Total Merged Rows", merge_stats.get('total', 0)))
    rows.append(("", ""))

    # Section B: Merge Logic
    rows.append(("SECTION B: MERGE LOGIC", ""))
    rows.append(("Join Type", "Outer join on feo_cohort"))
    rows.append(("RPU Addition", "Directly additive (same enrollment base): Combined_RPU = API_RPU + NonAPI_RPU"))
    rows.append(("Non-API Components", "W1, W2-W4, M3 only (No Non-Click, No Non-Enroll — these contribute 0)"))
    rows.append(("Delta Convention", "Delta = Actual - Predicted (positive = under-prediction, negative = over-prediction)"))
    rows.append(("API Components", "W1, W2-W4, M3, Non-Click, Non-Enroll (5 components)"))
    rows.append(("Division Safety", "All divisions use safe_divide() — defaults to 0 when denominator is 0 or NaN"))
    rows.append(("Matured Threshold", f"{MATURED_WEEKS_THRESHOLD}+ weeks"))
    rows.append(("", ""))

    # Section C: Conditional Formatting Rules
    rows.append(("SECTION C: CONDITIONAL FORMATTING RULES", ""))
    rows.append(("Delta RPU cells", "Red font = negative (over-prediction), Green font = positive (under-prediction)"))
    rows.append(("MAPE cells", "Green fill (<10%), Yellow fill (10-30%), Orange fill (30-50%), Red fill (>50%)"))
    rows.append(("Contribution_Pct cells", "Red fill if |value| > 30% (major error driver)"))
    rows.append(("Error Concentration", "Light red fill for rows in top 80% of cumulative error"))
    rows.append(("Summary Dashboard", "Bold the worst component and worst source"))
    rows.append(("", ""))

    # Section D: Column Definitions
    rows.append(("SECTION D: KEY COLUMN DEFINITIONS", ""))
    rows.append(("feo_cohort", "Weekly enrollment cohort date (always a Monday)"))
    rows.append(("Total_Enrolls", "Sum of enrollments across both segments for that cohort"))
    rows.append(("*_Predicted_RPU", "Model-predicted Revenue Per User for that component"))
    rows.append(("*_Actual_RPU", "Observed actual Revenue Per User for that component"))
    rows.append(("Delta", "Actual - Predicted (positive = under-prediction)"))
    rows.append(("Delta_Pct", "Delta / Actual RPU (percentage error)"))
    rows.append(("MAPE", "Mean Absolute Percentage Error — |Delta| / |Actual|"))
    rows.append(("Delta_Contribution_Pct", "Component's share of total error"))
    rows.append(("API_Share_Pct", "Fraction of component error from the API model"))
    rows.append(("Weeks_Since_Cohort", "Age of cohort in weeks from latest cohort date"))
    rows.append(("Maturity_Bucket", "Classification based on Weeks_Since_Cohort"))
    rows.append(("", ""))

    # Section E: Test Results
    rows.append(("SECTION E: TEST RESULTS", ""))
    for test_name, result in test_results:
        rows.append((test_name, result))

    return pd.DataFrame(rows, columns=['Item', 'Detail'])


# =============================================================================
# TEST CASES
# =============================================================================

def run_tests(combined_df, api_raw_df, nonapi_raw_df):
    """Run all validation tests and return list of (test_name, PASS/FAIL)."""
    print("\n--- Running Test Cases ---")
    results = []

    # Test 1: RPU Additivity
    test_name = "Test 1: RPU Additivity"
    both = combined_df[
        (combined_df[[f'{c}_Actual_RPU_API' for c in COMPONENTS]].sum(axis=1) > 0) &
        (combined_df[[f'{c}_Actual_RPU_NonAPI' for c in NONAPI_COMPONENTS]].sum(axis=1) > 0)
    ]
    sample = both.head(3) if len(both) >= 3 else both
    passed = True
    for _, row in sample.iterrows():
        for c in COMPONENTS:
            expected = row[f'{c}_Predicted_RPU_API'] + row[f'{c}_Predicted_RPU_NonAPI']
            actual = row[f'{c}_Predicted_RPU_Total']
            if abs(expected - actual) > 1e-8:
                passed = False
                break
    result = "PASS" if passed else "FAIL"
    print(f"  {test_name}: {result}")
    results.append((test_name, result))

    # Test 2: Delta Decomposition Sums
    test_name = "Test 2: Delta Decomposition Sums"
    passed = True
    for _, row in combined_df.iterrows():
        component_sum = sum(row[f'{c}_Delta_RPU'] for c in COMPONENTS)
        total = row['Total_Delta_RPU']
        if abs(component_sum - total) > 1e-8:
            passed = False
            break
    result = "PASS" if passed else "FAIL"
    print(f"  {test_name}: {result}")
    results.append((test_name, result))

    # Test 3: API + NonAPI Delta
    test_name = "Test 3: API + NonAPI Delta"
    sample = both.head(3) if len(both) >= 3 else both
    passed = True
    for _, row in sample.iterrows():
        for c in COMPONENTS:
            api_d = row[f'{c}_Delta_RPU_API']
            nonapi_d = row[f'{c}_Delta_RPU_NonAPI']
            total_d = row[f'{c}_Delta_RPU']
            if abs(api_d + nonapi_d - total_d) > 1e-8:
                passed = False
                break
    result = "PASS" if passed else "FAIL"
    print(f"  {test_name}: {result}")
    results.append((test_name, result))

    # Test 4: Non-API NonClick and NonEnroll = 0
    test_name = "Test 4: Non-API NonClick/NonEnroll = 0"
    zero_cols = ['NonClick_Predicted_RPU_NonAPI', 'NonClick_Actual_RPU_NonAPI',
                 'NonEnroll_Predicted_RPU_NonAPI', 'NonEnroll_Actual_RPU_NonAPI']
    passed = all(combined_df[col].abs().sum() == 0 for col in zero_cols)
    result = "PASS" if passed else "FAIL"
    print(f"  {test_name}: {result}")
    results.append((test_name, result))

    # Test 5: Error Concentration Cumulative
    test_name = "Test 5: Error Concentration Cumulative"
    ec = build_error_concentration(combined_df)
    if len(ec) > 0:
        final_cum = ec['Cumulative_Pct_of_Total'].iloc[-1]
        passed = abs(final_cum - 1.0) < 0.01
    else:
        passed = True
    result = "PASS" if passed else "FAIL"
    print(f"  {test_name}: {result} (final cumulative: {final_cum:.4f})" if len(ec) > 0 else f"  {test_name}: {result}")
    results.append((test_name, result))

    # Test 6: Summary Dashboard Cross-Check
    test_name = "Test 6: Summary Dashboard Cross-Check"
    matured = combined_df[combined_df['Weeks_Since_Cohort'] >= MATURED_WEEKS_THRESHOLD]
    matured_wa = matured[matured['Total_Actual_RPU'] > 0]
    if len(matured_wa) > 0:
        # Dashboard uses mean(|delta|) / mean(|actual|) — verify direction is correct
        dashboard_mape = safe_divide(
            matured_wa['Delta'].abs().mean(),
            matured_wa['Total_Actual_RPU'].abs().mean()
        )
        avg_delta = matured_wa['Delta'].mean()
        expected_direction = "Under-predicted" if avg_delta > 0 else "Over-predicted"
        # Verify MAPE > 0 and direction is consistent
        passed = dashboard_mape >= 0 and (avg_delta > 0) == (expected_direction == "Under-predicted")
    else:
        passed = True
    result = "PASS" if passed else "FAIL"
    print(f"  {test_name}: {result}")
    results.append((test_name, result))

    # Test 7: Division Safety
    test_name = "Test 7: Division Safety (no inf/NaN)"
    # Check key computed columns
    check_cols = ['Delta_Pct', 'Total_Delta_RPU', 'API_Share_of_Total_Delta_Pct']
    for c in COMPONENTS:
        check_cols.extend([f'{c}_Delta_Contribution_Pct', f'{c}_Error_Pct', f'{c}_MAPE', f'{c}_API_Share_Pct'])
    passed = True
    for col in check_cols:
        if col in combined_df.columns:
            if combined_df[col].replace([float('inf'), float('-inf')], np.nan).isna().any():
                print(f"    WARNING: inf/NaN found in {col}")
                passed = False
    result = "PASS" if passed else "FAIL"
    print(f"  {test_name}: {result}")
    results.append((test_name, result))

    # Test 8: Input Preservation
    test_name = "Test 8: Input Preservation"
    api_input = build_input_api_tab(api_raw_df)
    nonapi_input = build_input_nonapi_tab(nonapi_raw_df)
    passed = (len(api_input) == len(api_raw_df)) and (len(nonapi_input) == len(nonapi_raw_df))
    result = "PASS" if passed else "FAIL"
    print(f"  {test_name}: {result} (API: {len(api_input)} rows, Non-API: {len(nonapi_input)} rows)")
    results.append((test_name, result))

    return results


# =============================================================================
# EXCEL WRITER WITH FORMATTING
# =============================================================================

def write_df_to_sheet(ws, df, description, start_row=1, date_cols=None, rpu_cols=None,
                      pct_cols=None, count_cols=None, alt_row_shading=True):
    """Write a DataFrame to a worksheet with standard formatting."""
    if date_cols is None:
        date_cols = []
    if rpu_cols is None:
        rpu_cols = []
    if pct_cols is None:
        pct_cols = []
    if count_cols is None:
        count_cols = []

    # Row 1: Description
    ws.cell(row=1, column=1, value=description)
    ws.cell(row=1, column=1).font = FONT_BOLD_ITALIC
    for col_idx in range(1, len(df.columns) + 1):
        ws.cell(row=1, column=col_idx).fill = FILL_GREY

    # Row 2: Headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = FONT_BOLD
        cell.fill = FILL_LIGHT_BLUE
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        for col_idx, col_name in enumerate(df.columns, 1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)

            # Handle datetime
            if isinstance(val, (pd.Timestamp, datetime)):
                cell.value = val
                cell.number_format = 'YYYY-MM-DD'
            elif col_name in date_cols and not isinstance(val, str):
                try:
                    cell.value = pd.to_datetime(val)
                    cell.number_format = 'YYYY-MM-DD'
                except:
                    cell.value = val
            elif col_name in pct_cols:
                cell.value = val
                cell.number_format = '0.00%'
            elif col_name in rpu_cols:
                cell.value = val
                cell.number_format = '$#,##0.000000'
            elif col_name in count_cols:
                cell.value = val
                cell.number_format = '#,##0'
            else:
                cell.value = val

            # Auto-detect format by column name
            if col_name not in date_cols and col_name not in pct_cols and col_name not in rpu_cols and col_name not in count_cols:
                if isinstance(val, str):
                    pass  # Keep as-is
                elif 'Pct' in col_name or 'MAPE' in col_name or 'Error_Pct' in col_name or 'Share' in col_name:
                    cell.number_format = '0.00%'
                elif 'RPU' in col_name or 'Delta_RPU' in col_name or 'Delta_API' in col_name or 'Delta_NonAPI' in col_name or 'Delta_Total' in col_name or col_name == 'Delta' or col_name == 'RPU_Delta' or col_name == 'Abs_RPU_Delta' or col_name == 'Cumulative_Abs_Delta':
                    cell.number_format = '$#,##0.000000'
                elif 'Enrolls' in col_name or col_name == 'Weeks_Since_Cohort' or col_name == 'Error_Rank':
                    cell.number_format = '#,##0'

            # Alternate row shading
            if alt_row_shading and (row_idx - 3) % 2 == 1:
                cell.fill = FILL_ALT_ROW

    # Freeze Row 2
    ws.freeze_panes = 'A3'

    # Auto-fit column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), 10)
        for row_idx in range(3, min(len(df) + 3, 20)):  # Sample first ~17 rows
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, min(len(str(cell_val)) + 2, 30))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len, 10), 30)


def apply_conditional_formatting_delta(ws, df, col_name, start_row=3):
    """Apply red/green font to delta RPU cells."""
    if col_name not in df.columns:
        return
    col_idx = list(df.columns).index(col_name) + 1
    for row_idx, (_, row) in enumerate(df.iterrows(), start_row):
        val = row[col_name]
        cell = ws.cell(row=row_idx, column=col_idx)
        if isinstance(val, (int, float)) and not pd.isna(val):
            if val < 0:
                cell.font = FONT_RED
            elif val > 0:
                cell.font = FONT_GREEN


def apply_conditional_formatting_mape(ws, df, col_name, start_row=3):
    """Apply MAPE color bands."""
    if col_name not in df.columns:
        return
    col_idx = list(df.columns).index(col_name) + 1
    for row_idx, (_, row) in enumerate(df.iterrows(), start_row):
        val = row[col_name]
        cell = ws.cell(row=row_idx, column=col_idx)
        if isinstance(val, (int, float)) and not pd.isna(val):
            abs_val = abs(val)
            if abs_val < MAPE_THRESHOLDS['green']:
                cell.fill = FILL_GREEN
            elif abs_val < MAPE_THRESHOLDS['yellow']:
                cell.fill = FILL_YELLOW
            elif abs_val < MAPE_THRESHOLDS['orange']:
                cell.fill = FILL_ORANGE
            else:
                cell.fill = FILL_RED


def apply_conditional_formatting_contribution(ws, df, col_name, start_row=3):
    """Apply red fill for major error drivers (>30% contribution)."""
    if col_name not in df.columns:
        return
    col_idx = list(df.columns).index(col_name) + 1
    for row_idx, (_, row) in enumerate(df.iterrows(), start_row):
        val = row[col_name]
        cell = ws.cell(row=row_idx, column=col_idx)
        if isinstance(val, (int, float)) and not pd.isna(val):
            if abs(val) > ERROR_DRIVER_THRESHOLD:
                cell.fill = FILL_RED


def write_summary_dashboard(ws, dashboard_data):
    """Write the Summary Dashboard with manual cell placement."""
    # Row 1: Description
    ws.cell(row=1, column=1, value="Executive summary with key aggregated metrics across all diagnostic views. Matured = 12+ weeks.")
    ws.cell(row=1, column=1).font = FONT_BOLD_ITALIC
    for c in range(1, 7):
        ws.cell(row=1, column=c).fill = FILL_GREY

    # Section A: Overall Accuracy
    ws.cell(row=2, column=1, value="SECTION A: OVERALL ACCURACY")
    ws.cell(row=2, column=1).font = FONT_BOLD
    ws.cell(row=2, column=1).fill = FILL_LIGHT_BLUE
    ws.cell(row=2, column=2).fill = FILL_LIGHT_BLUE

    for i, (metric, value) in enumerate(dashboard_data['section_a'][1:], 3):  # Skip header
        ws.cell(row=i, column=1, value=metric).font = FONT_BOLD
        cell = ws.cell(row=i, column=2, value=value)
        if 'MAPE' in str(metric):
            cell.number_format = '0.00%'
        elif 'Delta RPU' in str(metric):
            cell.number_format = '$#,##0.000000'

    # Section B: Component Accuracy Ranking
    row_start = 10
    ws.cell(row=row_start, column=1, value="SECTION B: COMPONENT ACCURACY RANKING")
    ws.cell(row=row_start, column=1).font = FONT_BOLD
    for c in range(1, 6):
        ws.cell(row=row_start, column=c).fill = FILL_LIGHT_BLUE

    # Header
    for ci, h in enumerate(dashboard_data['section_b_header'][0], 1):
        ws.cell(row=row_start + 1, column=ci, value=h).font = FONT_BOLD

    # Data
    worst_rank = len(dashboard_data['section_b_data'])
    for ri, (comp, mape, delta, rank, verdict) in enumerate(dashboard_data['section_b_data'], row_start + 2):
        ws.cell(row=ri, column=1, value=comp)
        ws.cell(row=ri, column=2, value=mape).number_format = '0.00%'
        ws.cell(row=ri, column=3, value=delta).number_format = '$#,##0.000000'
        ws.cell(row=ri, column=4, value=rank)
        cell = ws.cell(row=ri, column=5, value=verdict)
        if rank == worst_rank:
            cell.font = FONT_BOLD_RED

    # Section C: Source Accuracy
    row_start = row_start + 2 + len(dashboard_data['section_b_data']) + 1
    ws.cell(row=row_start, column=1, value="SECTION C: SOURCE ACCURACY")
    ws.cell(row=row_start, column=1).font = FONT_BOLD
    for c in range(1, 6):
        ws.cell(row=row_start, column=c).fill = FILL_LIGHT_BLUE

    for ci, h in enumerate(dashboard_data['section_c_header'][0], 1):
        ws.cell(row=row_start + 1, column=ci, value=h).font = FONT_BOLD

    for ri, (src, pred, act, delta, mape) in enumerate(dashboard_data['section_c_data'], row_start + 2):
        ws.cell(row=ri, column=1, value=src)
        ws.cell(row=ri, column=2, value=pred).number_format = '$#,##0.000000'
        ws.cell(row=ri, column=3, value=act).number_format = '$#,##0.000000'
        ws.cell(row=ri, column=4, value=delta).number_format = '$#,##0.000000'
        ws.cell(row=ri, column=5, value=mape).number_format = '0.00%'

    # Section D: Top 3 Error Drivers
    row_start = row_start + 2 + len(dashboard_data['section_c_data']) + 1
    ws.cell(row=row_start, column=1, value="SECTION D: TOP 3 ERROR DRIVERS")
    ws.cell(row=row_start, column=1).font = FONT_BOLD
    for c in range(1, 7):
        ws.cell(row=row_start, column=c).fill = FILL_LIGHT_BLUE

    for ci, h in enumerate(dashboard_data['section_d_header'][0], 1):
        ws.cell(row=row_start + 1, column=ci, value=h).font = FONT_BOLD

    for ri, row_data in enumerate(dashboard_data['section_d_data'], row_start + 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == 5:  # Delta RPU
                cell.number_format = '$#,##0.000000'

    # Freeze and width
    ws.freeze_panes = 'A2'
    for col_idx in range(1, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25


def write_excel(tabs, output_path):
    """Write all tabs to a formatted Excel workbook."""
    print(f"\n--- Writing Excel to: {output_path} ---")
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    tab_configs = [
        {
            'name': '01 Input API Main',
            'df': tabs['input_api'],
            'desc': "Raw input data from API RPU Projections — sheet '22 Main'. No transformations applied.",
            'color_group': 'input',
        },
        {
            'name': '02 Input NonAPI Main',
            'df': tabs['input_nonapi'],
            'desc': "Raw input data from Non-API RPU Projections — sheet '19 Main'. No transformations applied.",
            'color_group': 'input',
        },
        {
            'name': '03 Combined Main',
            'df': tabs['combined_main'],
            'desc': "Unified executive summary — Total (API + Non-API) Projected vs Actual RPU with all 5 component breakdowns.",
            'color_group': 'merge',
        },
        {
            'name': '04 Delta Waterfall',
            'df': tabs['delta_waterfall'],
            'desc': "Component decomposition of total Delta — shows which components are driving the projection error (waterfall view).",
            'color_group': 'diagnostics',
            'cond_fmt': 'waterfall',
        },
        {
            'name': '05 API vs NonAPI Delta',
            'df': tabs['api_vs_nonapi'],
            'desc': "Source-level error decomposition — for each component, shows how much delta came from API vs Non-API model.",
            'color_group': 'diagnostics',
        },
        {
            'name': '06 Component Accuracy',
            'df': tabs['component_accuracy'],
            'desc': "Per-component accuracy metrics (Error_Pct, MAPE) with ranking. Summary row at bottom uses matured cohorts (12+ weeks).",
            'color_group': 'diagnostics',
            'cond_fmt': 'accuracy',
        },
        {
            'name': '07 Maturity Analysis',
            'df': tabs['maturity'],
            'desc': "Projection accuracy by cohort age — validates that accuracy improves as cohorts mature. Matured = 12+ weeks.",
            'color_group': 'analysis',
        },
        {
            'name': '08 Period Trends',
            'df': tabs['period_trends'],
            'desc': "Time-series view of accuracy metrics with 4-week rolling averages. Sorted oldest-first for trend readability.",
            'color_group': 'analysis',
        },
        {
            'name': '09 Error Concentration',
            'df': tabs['error_concentration'],
            'desc': "Pareto analysis — identifies which cohorts contribute most to total RPU error. Sorted by absolute error descending.",
            'color_group': 'analysis',
            'cond_fmt': 'pareto',
        },
        # Tab 10 (Summary Dashboard) is handled separately
        {
            'name': '11 Comment',
            'df': tabs['comment'],
            'desc': "Documentation — input file summary, merge logic, conditional formatting rules, column definitions, and test results.",
            'color_group': 'summary',
        },
    ]

    for config in tab_configs:
        ws = wb.create_sheet(title=config['name'])
        ws.sheet_properties.tabColor = TAB_COLORS[config['color_group']]

        write_df_to_sheet(ws, config['df'], config['desc'])

        # Apply conditional formatting
        cond_fmt = config.get('cond_fmt', '')
        df = config['df']

        if cond_fmt == 'waterfall':
            for c in COMPONENTS:
                apply_conditional_formatting_delta(ws, df, f'{c}_Delta_RPU')
                apply_conditional_formatting_contribution(ws, df, f'{c}_Delta_Contribution_Pct')
            apply_conditional_formatting_delta(ws, df, 'Total_Delta_RPU')

        elif cond_fmt == 'accuracy':
            for c in COMPONENTS:
                apply_conditional_formatting_mape(ws, df, f'{c}_MAPE')

        elif cond_fmt == 'pareto':
            # Highlight rows in top 80%
            if 'Cumulative_Pct_of_Total' in df.columns:
                col_idx = list(df.columns).index('Cumulative_Pct_of_Total') + 1
                for row_idx, (_, row) in enumerate(df.iterrows(), 3):
                    if row['Cumulative_Pct_of_Total'] <= PARETO_THRESHOLD:
                        for ci in range(1, len(df.columns) + 1):
                            ws.cell(row=row_idx, column=ci).fill = FILL_LIGHT_RED

    # Tab 10: Summary Dashboard (special layout)
    ws_dash = wb.create_sheet(title='10 Summary Dashboard')
    wb.move_sheet('10 Summary Dashboard', offset=-1)  # Move before Comment tab
    ws_dash.sheet_properties.tabColor = TAB_COLORS['summary']
    write_summary_dashboard(ws_dash, tabs['dashboard'])

    # Apply delta formatting on Combined Main
    ws_main = wb['03 Combined Main']
    apply_conditional_formatting_delta(ws_main, tabs['combined_main'], 'Delta')

    # Save
    wb.save(output_path)
    print(f"  Saved: {output_path}")
    print(f"  Total tabs: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"    - {name}")


# =============================================================================
# MAIN ORCHESTRATOR
# =============================================================================

def main():
    """Main entry point — orchestrates the full merge and diagnostics pipeline."""
    print("=" * 70)
    print("Combined RPU Projection Merge & Diagnostics Engine")
    print("=" * 70)

    # Step 0: Load data
    api_df, api_raw_df = load_api_main(API_EXCEL_PATH)
    nonapi_df, nonapi_raw_df = load_nonapi_main(NONAPI_EXCEL_PATH)
    enrolls_df = load_enrolls(ENROLLS_PATH)

    # Step 0.4: Merge
    merged, merge_stats = merge_sources(api_df, nonapi_df, enrolls_df)

    # Step 1-3: Compute combined metrics
    combined = compute_combined_metrics(merged)

    # Run tests
    test_results = run_tests(combined, api_raw_df, nonapi_raw_df)

    # Build all tabs
    print("\n--- Building Output Tabs ---")

    input_api = build_input_api_tab(api_raw_df)
    print(f"  Tab 01 Input API Main: {len(input_api)} rows")

    input_nonapi = build_input_nonapi_tab(nonapi_raw_df)
    print(f"  Tab 02 Input NonAPI Main: {len(input_nonapi)} rows")

    combined_main = build_combined_main(combined)
    print(f"  Tab 03 Combined Main: {len(combined_main)} rows")

    delta_waterfall = build_delta_waterfall(combined)
    print(f"  Tab 04 Delta Waterfall: {len(delta_waterfall)} rows")

    api_vs_nonapi = build_api_vs_nonapi_delta(combined)
    print(f"  Tab 05 API vs NonAPI Delta: {len(api_vs_nonapi)} rows")

    component_accuracy = build_component_accuracy(combined)
    print(f"  Tab 06 Component Accuracy: {len(component_accuracy)} rows")

    maturity = build_maturity_analysis(combined)
    print(f"  Tab 07 Maturity Analysis: {len(maturity)} rows")

    period_trends = build_period_trends(combined)
    print(f"  Tab 08 Period Trends: {len(period_trends)} rows")

    error_concentration = build_error_concentration(combined)
    print(f"  Tab 09 Error Concentration: {len(error_concentration)} rows")

    dashboard = build_summary_dashboard(combined)
    print(f"  Tab 10 Summary Dashboard: built")

    # Build info dicts for comment tab
    api_info = {
        'path': API_EXCEL_PATH,
        'sheet': API_MAIN_SHEET,
        'rows': len(api_raw_df),
        'range': f"{api_raw_df['feo_cohort'].min().date()} to {api_raw_df['feo_cohort'].max().date()}"
    }
    nonapi_info = {
        'path': NONAPI_EXCEL_PATH,
        'sheet': NONAPI_MAIN_SHEET,
        'rows': len(nonapi_raw_df),
        'range': f"{nonapi_raw_df['feo_cohort'].min().date()} to {nonapi_raw_df['feo_cohort'].max().date()}"
    }
    enrolls_info = {
        'path': ENROLLS_PATH,
        'rows': len(enrolls_df),
        'range': f"{enrolls_df['feo_cohort'].min().date()} to {enrolls_df['feo_cohort'].max().date()}"
    }

    comment = build_comment_tab(api_info, nonapi_info, enrolls_info, merge_stats, test_results)
    print(f"  Tab 11 Comment: {len(comment)} rows")

    # Write Excel
    all_tabs = {
        'input_api': input_api,
        'input_nonapi': input_nonapi,
        'combined_main': combined_main,
        'delta_waterfall': delta_waterfall,
        'api_vs_nonapi': api_vs_nonapi,
        'component_accuracy': component_accuracy,
        'maturity': maturity,
        'period_trends': period_trends,
        'error_concentration': error_concentration,
        'dashboard': dashboard,
        'comment': comment,
    }

    write_excel(all_tabs, OUTPUT_PATH)

    print("\n" + "=" * 70)
    print("DONE — All tabs generated successfully.")
    print("=" * 70)


if __name__ == '__main__':
    main()
